
import json
import math
import numpy
import os
import pytz

from apps.auctions.models import (
    Auction,
    AuctionSupplier,
    AuctionItem,
    AuctionItemSupplier,
    AuctionBid,
    AuctionTypeTrafficLight,
    AuctionTypePrices,
    AuctionTypeRanking,
    AuctionTypeSealedBid,
    AuctionTypeDutch,
    AuctionTypeJapanese,
    AuctionOtherRequirement,
    AuctionGeneralTermCondition,
)
from apps.auctions.serializers import (
    AuctionSerializer,
    AuctionSupplierSerializer,
    AuctionBySupplierSerializer,
    AuctionDetailSerializer,
    AuctionDetailBySupplierSerializer,
    AuctionItemSerializer,
    AuctionItemSupplierSerializer,
    AuctionSupplierResultSerializer,
    AuctionResultsSerializer,
    AuctionTypeTrafficLightSerializer,
    AuctionTypeRankingSerializer,
    AuctionTypePricesSerializer,
    AuctionTypeSealedBidSerializer,
    AuctionTypeDutchSerializer,
    AuctionTypeJapaneseSerializer,
    GetAuctionSerializer,
    AuctionOtherRequirementSerializer,
    AuctionGeneralTermConditionSerializer,
    AuctionReportExportSerializer,
)
from apps.invoices.views import invoice_generate
from apps.master_data.models import (
    EmailTemplates,
    Coupon,
    PaymentTermTranslation,
    DeliveryTermTranslation,
    TechnicalWeightingTranslation,
    ContractTypeTranslation,
    CurrencyTranslation,
    EmailTemplatesTranslation,
)
from apps.payment.models import History, PaymentAuction
from apps.sale_schema.models import AuctionFee, PlatformFee
from apps.users.models import Buyer, BuyerSubAccounts
from datetime import datetime, timedelta

from django_filters import rest_framework as filters
from django_filters import FilterSet
from django_filters.rest_framework import DjangoFilterBackend
from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.mail import send_mail
from django.db import transaction
from django.db.models import Max
from django.http import HttpResponse, QueryDict
from django.template import Template, Context
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

from pathlib import Path

from rest_framework import status, generics, permissions, parsers, serializers
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.filters import OrderingFilter

User = get_user_model()

class CreateAuction(APIView):

    permission_classes = [permissions.IsAuthenticated]
    parsers_classes = [parsers.FormParser, parsers.MultiPartParser]

    """
    Creates auction.
    """

    def post(self, request, format='json'):

        auction_count = Auction.objects.filter(auction_next_round__isnull=True).count() + 1
        item_code = '60' + str(auction_count).zfill(4)
        auction_next_round = request.data.get('auction_next_round')
        if auction_next_round is not None:
            auction = Auction.objects.filter(id=auction_next_round).first()
            AuctionSupplier.objects.filter(auction_id=auction_next_round).exclude(supplier_status=7).update(supplier_status=11)

            item_code = float("{0:.2f}".format(float(auction.item_code) + 0.1))

        if isinstance(request.data, QueryDict):
            data = {
                **request.data.dict(),
                'item_code': item_code,
            }
        else:
            data = {
                **request.data,
                'item_code': item_code,
            }
        other_requirements_input = request.data.getlist('other_requirements')
        general_term_conditions_input = request.data.getlist('general_term_conditions')

        if data.get('duplicated_auction_id') is not None:
            auction_copy = Auction.objects.filter(pk=data.get('duplicated_auction_id')).first()
            if len(general_term_conditions_input) == 0 and bool(int(data.get('is_general_term_condition_deleted'))) == False:
                for i in AuctionGeneralTermCondition.objects.filter(auction=auction_copy):
                    general_term_conditions_input.append(i.general_term_condition)
            if len(other_requirements_input) == 0 and bool(int(data.get('is_other_requirement_deleted'))) == False:
                for i in AuctionOtherRequirement.objects.filter(auction=auction_copy):
                    other_requirements_input.append(i.other_requirement)

        user_id = request.data.get('user_id')
        user = User.objects.filter(id=user_id).first()
        list_id = []

        if user.company_position == 1:
            buyer = Buyer.objects.get(user_id=user.id)
            buyer_auction = buyer.profile_features.no_eauction_year
            list_sub_account = BuyerSubAccounts.objects.filter(buyer=buyer)
            for ob in list_sub_account:
                list_id.append(ob.user_id)
            list_id.append(buyer.user_id)

        else:
            buyer_sub_accounts = BuyerSubAccounts.objects.get(user_id=user.id)
            buyer = Buyer.objects.get(id=buyer_sub_accounts.buyer_id)
            buyer_auction = buyer.profile_features.no_eauction_year
            list_sub_account = BuyerSubAccounts.objects.filter(buyer_id=buyer.id)
            for ob in list_sub_account:
                list_id.append(ob.user_id)
            list_id.append(buyer.user_id)

        count = Auction.objects.filter(user_id__in=list_id).count()
        if count < buyer_auction:
            try:
                auctionSerializer = AuctionSerializer(data=data)

                auctionSerializer.is_valid(raise_exception=True)

                auction = auctionSerializer.save()

                for other_requirement in other_requirements_input:
                    auctionOtherRequirementSerializer = AuctionOtherRequirementSerializer(
                        data={'other_requirement': other_requirement, 'auction_id': auction.id}
                    )
                    auctionOtherRequirementSerializer.is_valid(raise_exception=True)
                    auctionOtherRequirementSerializer.save()

                for general_term_condition in general_term_conditions_input:
                    auctionGeneralTermConditionSerializer = AuctionGeneralTermConditionSerializer(
                        data={'general_term_condition': general_term_condition, 'auction_id': auction.id}
                    )
                    auctionGeneralTermConditionSerializer.is_valid(raise_exception=True)
                    auctionGeneralTermConditionSerializer.save()

                data['options'] = json.loads(request.data.get('options'))
                data['items'] = json.loads(request.data.get('items'))

                auctionTypeData = {**data['options'], 'auction_id': auction.id}

                # Count suppliers of auction
                auction_items = data['items']
                number_of_items = len(auction_items)

                s = 0
                for item in auction_items:
                    auction_suppliers = item['suppliers']
                    number_of_suppliers = len(auction_suppliers)
                    s = s + number_of_suppliers

                suppliers_of_auction = s / number_of_items

                # -----------------------------------------------------------------------
                # -------About-max(starting)-price--------------------------------------
                # -------Setting-max-price-of-Traffic-Light-or-Ranking-or-Prices-are-same-----------------

                if auction.auction_type1_id == 1 or auction.auction_type1_id == 3 or auction.auction_type1_id == 4:
                    # --Case: False, Null, Null -----------------------------------
                    if auctionTypeData['max_price_settings'] is False:
                        auctionTypeData['max_price_type'] = None
                        auctionTypeData['individual_or_max_price'] = None
                        auctionTypeData['entire_auction'] = None
                    elif auctionTypeData['max_price_settings'] is True:
                        if auctionTypeData['max_price_type'] is None:
                            return Response({'error': 'Please choose type of max price when selecting True for setting max price.'})

                        if auctionTypeData['individual_or_max_price'] is None:
                            return Response({'error': 'Please check individual or maximum price'})

                        if auctionTypeData['max_price_type'] == 1 and auctionTypeData['individual_or_max_price'] == 2:
                            if auctionTypeData['entire_auction'] is None:
                                return Response({'error': 'Please fill max price for entire Auction.'})
                            else:
                                pass
                        else:
                            auctionTypeData['entire_auction'] = None

                # -------------------------------------------------------------------------------------------------------------------------------

                # --------------------------------------------------------------------------------------------------

                if auction.auction_type1_id == 4:  # Prices
                    # Check of bid mornitoring - views disabled and extension--------------------------------------

                    if auctionTypeData['type_of_bid_monitoring6'] is False:
                        auctionTypeData['views_disabled'] = None
                        auctionTypeData['minutes'] = None
                    elif auctionTypeData['type_of_bid_monitoring6'] is True:
                        if auctionTypeData['views_disabled'] == 1:
                            auctionTypeData['minutes'] = None
                            auctionTypeData['type_of_bid_monitoring1'] = False

                        if auctionTypeData['views_disabled'] == 2 and auctionTypeData['minutes'] is None:
                            return Response({'error': 'Please fill the rest time you want to the information of suppliers be disabled.'})

                    if auctionTypeData['auction_extension'] == 1:
                        auctionTypeData['auction_extension_trigger'] = None
                        auctionTypeData['number_of_rankings'] = None
                        auctionTypeData['frequency'] = None
                        auctionTypeData['trigger_time'] = None
                        auctionTypeData['prolongation_by'] = None
                    else:
                        if auctionTypeData['type_of_bidding'] == 2:
                            auctionTypeData['auction_extension_trigger'] = 1
                            auctionTypeData['number_of_rankings'] = None
                            if (
                                auctionTypeData['frequency'] is None
                                or auctionTypeData['prolongation_by'] is None
                                or auctionTypeData['trigger_time'] is None
                            ):
                                return Response({'error': 'You must type the required field(s).'})
                        if auctionTypeData['type_of_bidding'] == 1:
                            if auctionTypeData['auction_extension_trigger'] is None:
                                return Response({'error': 'You must select the way how to trigger when chosing Automatic extension.'})

                            if auctionTypeData['auction_extension_trigger'] == 2:
                                if auctionTypeData['number_of_rankings'] is None:
                                    return Response({'error': 'You must type the number of ranking when chosing the second way trigger.'})
                                if auctionTypeData['number_of_rankings'] > suppliers_of_auction:
                                    return Response({'error': 'You must type a number lower than or equal to the number of suppliers.'})

                            if auctionTypeData['frequency'] is None or auctionTypeData['prolongation_by'] is None:
                                return Response({'error': 'You must type the required field(s).'})

                            if auctionTypeData['auction_extension_trigger'] != 2:
                                auctionTypeData['number_of_rankings'] = None

                            if auctionTypeData['auction_extension_trigger'] == 3:
                                auctionTypeData['trigger_time'] = None
                            else:
                                if auctionTypeData['trigger_time'] is None:
                                    return Response({'error': 'You must type the required trigger time.'})

                # --------------------------------------------------------------------------------------------------

                if auction.auction_type1_id == 1:  # Traffic Light
                    # ------Check-case-of-bid-mornitoring----------------------------------------
                    # ------Check-3rd-then check-or-not-4th----------------------------------------
                    # if auctionTypeData['type_of_bid_monitoring3'] is False

                    # -------Check-6th-then-disabled-all-rest---------------------------------------
                    if auctionTypeData['type_of_bid_monitoring6'] is False:
                        auctionTypeData['views_disabled'] = None
                        auctionTypeData['minutes'] = None
                    elif auctionTypeData['type_of_bid_monitoring6'] is True:
                        if auctionTypeData['views_disabled'] == 1:
                            auctionTypeData['minutes'] = None
                            auctionTypeData['type_of_bid_monitoring1'] = False
                            auctionTypeData['type_of_bid_monitoring2'] = False
                            auctionTypeData['type_of_bid_monitoring3'] = False

                        if auctionTypeData['views_disabled'] == 2 and auctionTypeData['minutes'] is None:
                            return Response({'error': 'Please fill the time you want to the information of suppliers be disabled.'})
                # Ranking
                if auction.auction_type1_id == 3:
                    # ------Check-case-of-bid-mornitoring----------------------------------------
                    # ------Check-3rd-then check-or-not-4th----------------------------------------
                    # if auctionTypeData['type_of_bid_monitoring3'] is False

                    # -------Check-6th-then-disabled-all-rest---------------------------------------
                    if auctionTypeData['type_of_bid_monitoring6'] is False:
                        auctionTypeData['views_disabled'] = None
                        auctionTypeData['minutes'] = None
                    elif auctionTypeData['type_of_bid_monitoring6'] is True:
                        if auctionTypeData['views_disabled'] == 1:
                            auctionTypeData['minutes'] = None
                            auctionTypeData['type_of_bid_monitoring1'] = False
                            auctionTypeData['type_of_bid_monitoring3'] = False

                        if auctionTypeData['views_disabled'] == 2 and auctionTypeData['minutes'] is None:
                            return Response({'error': 'Please fill the time you want to the information of suppliers be disabled.'})

                    # -------------------------------------------------------------------------------------------------------------
                    # ---------Type of bidding affect directly to extension trigger ------------------------------------------------

                    if auctionTypeData['type_of_bid_monitoring3'] is False:
                        if auctionTypeData['type_of_bidding'] == 2:
                            return Response({'error': 'Default: Bidder sees the best bid '})
                    if auctionTypeData['auction_extension'] == 1:
                        auctionTypeData['auction_extension_trigger'] = None
                        auctionTypeData['number_of_rankings'] = None
                        auctionTypeData['frequency'] = None
                        auctionTypeData['trigger_time'] = None
                        auctionTypeData['prolongation_by'] = None
                    else:
                        if auctionTypeData['type_of_bidding'] == 2:
                            auctionTypeData['auction_extension_trigger'] = 1
                            auctionTypeData['number_of_rankings'] = None
                            if (
                                auctionTypeData['frequency'] is None
                                or auctionTypeData['prolongation_by'] is None
                                or auctionTypeData['trigger_time'] is None
                            ):
                                return Response({'error': 'You must type the required field(s).'})
                        if auctionTypeData['type_of_bidding'] == 1:
                            if auctionTypeData['auction_extension_trigger'] is None:
                                return Response({'error': 'You must select the way how to trigger when chosing Automatic extension.'})

                            if auctionTypeData['auction_extension_trigger'] == 2:
                                if auctionTypeData['number_of_rankings'] is None:
                                    return Response({'error': 'You must type the number of ranking when chosing the second way trigger.'})
                                if auctionTypeData['number_of_rankings'] > suppliers_of_auction:
                                    return Response({'error': 'You must type a number lower than or equal to the number of suppliers.'})

                            if auctionTypeData['frequency'] is None or auctionTypeData['prolongation_by'] is None:
                                return Response({'error': 'You must type the required field(s).'})

                            if auctionTypeData['auction_extension_trigger'] != 2:
                                auctionTypeData['number_of_rankings'] = None

                            if auctionTypeData['auction_extension_trigger'] == 3:
                                auctionTypeData['trigger_time'] = None
                            else:
                                if auctionTypeData['trigger_time'] is None:
                                    return Response({'error': 'You must type the required trigger time.'})
                # ----------Check condition of Dutch Auction--------------------
                if auction.auction_type1_id == 5:
                    if auctionTypeData['initial_price'] >= auctionTypeData['end_price']:
                        return Response({'error': 'Initial price must be lower than end price.'})

                    if (
                        auctionTypeData['initial_price'] is None
                        or auctionTypeData['end_price'] is None
                        or auctionTypeData['price_step'] is None
                        or auctionTypeData['price_validity'] is None
                    ):
                        return Response({'error': 'You must type the required field(s).'})

                    if (auctionTypeData['initial_price'] + auctionTypeData['price_step']) > auctionTypeData['end_price']:
                        return Response({'error': 'You must type valid price step.'})

                    round_auction = (auctionTypeData['end_price'] - auctionTypeData['initial_price']) / auctionTypeData['price_step']
                    round_auction = math.ceil(round_auction) + 1

                # ----------Check condition of Japanese Auction--------------------
                if auction.auction_type1_id == 6:
                    if auctionTypeData['initial_price'] <= auctionTypeData['end_price']:
                        return Response({'error': 'Initial price must be greater than end price.'})

                    if (
                        auctionTypeData['initial_price'] is None
                        or auctionTypeData['end_price'] is None
                        or auctionTypeData['price_step'] is None
                        or auctionTypeData['price_validity'] is None
                    ):
                        return Response({'error': 'You must type the required field(s).'})

                    if (auctionTypeData['initial_price'] - auctionTypeData['price_step']) < auctionTypeData['end_price']:
                        return Response({'error': 'You must type valid price step.'})

                    round_auction = (auctionTypeData['initial_price'] - auctionTypeData['end_price']) / auctionTypeData['price_step']
                    round_auction = math.ceil(round_auction) + 1

                # Start to save E_Auction Type
                if auction.auction_type1_id == 1:  # Traffic Light
                    auctionTypeSerializer = AuctionTypeTrafficLightSerializer(data=auctionTypeData)
                elif auction.auction_type1_id == 2:  # Sealed Bid
                    auctionTypeSerializer = AuctionTypeSealedBidSerializer(data=auctionTypeData)
                elif auction.auction_type1_id == 3:  # Ranking
                    auctionTypeSerializer = AuctionTypeRankingSerializer(data=auctionTypeData)
                elif auction.auction_type1_id == 4:  # Prices
                    auctionTypeSerializer = AuctionTypePricesSerializer(data=auctionTypeData)
                elif auction.auction_type1_id == 5:
                    auctionTypeSerializer = AuctionTypeDutchSerializer(data={**auctionTypeData, 'round_auction': round_auction})
                elif auction.auction_type1_id == 6:
                    auctionTypeSerializer = AuctionTypeJapaneseSerializer(data={**auctionTypeData, 'round_auction': round_auction})

                auctionTypeSerializer.is_valid(raise_exception=True)
                auctionType = auctionTypeSerializer.save()

                # ---Case: entire Auction--------Calculate the average of all items and compare to max price of all items----------------
                # ---Case: True, 1, 2---------------------database  updated-------------------------
                if (
                    auction.auction_type1_id == 1 or auction.auction_type1_id == 3 or auction.auction_type1_id == 4
                ):  # add this line as Sealed bid dooes not have fields about max price
                    if auctionType.max_price_settings is True and auctionType.max_price_type == 1 and auctionType.individual_or_max_price == 2:
                        average_price_total = 0
                        price_item = 0
                        for item in data['items']:
                            for supplier in item['suppliers']:
                                price_item = price_item + supplier['price']
                            average_price_total = price_item / len(item['suppliers'])

                        for item in data['items']:
                            item['max_price'] = None

                        max_price_all_items = auctionTypeData['entire_auction']

                        percent = 0.8
                        bounded_max_price = percent * average_price_total
                        error_max_and_average_total = abs(max_price_all_items - average_price_total)

                        if error_max_and_average_total > bounded_max_price:
                            return Response(
                                {
                                    'Note': 'You choose max price for entire Auction.',
                                    'error': 'Please fill in max price such that value of all items between 0,2 and 1,8 average of all items.',
                                }
                            )

                # ----------------------------------------------------------------------------
                user_buyer = User.objects.select_related('buyer').get(id=auction.user_id)

                # Start to save items
                for item in data['items']:
                    auctionItemData = {**item, 'auction_id': auction.id}
                    if auction.auction_type1_id == 1:  # Traffic Light
                        if (
                            auctionItemData['target_price'] is None
                            or auctionItemData['price_yellow'] is None
                            or auctionItemData['minimum_bid_step'] is None
                            or auctionItemData['maximum_bid_step'] is None
                        ):
                            return Response(
                                {
                                    'error': 'Please check price level for yellow, target price, minimum bid step and maximum bid step, maybe one of these fields be incomplete'
                                }
                            )

                        if (
                            auctionType.max_price_settings is True
                            and auctionType.max_price_type == 2
                            and auctionType.individual_or_max_price == 2
                            and auctionItemData['max_price'] is None
                        ):
                            return Response({'error': 'Please fill value of max price for all items as chosing setting True'})

                        if auctionType.max_price_settings is False:
                            auctionItemData['max_price'] = None

                        if auctionType.individual_or_max_price == 1:
                            auctionItemData['max_price'] = None

                    if auction.auction_type1_id == 3 or auction.auction_type1_id == 4:  # Ranking-Prices
                        auctionItemData['price_yellow'] = None

                        if auctionItemData['minimum_bid_step'] is None or auctionItemData['maximum_bid_step'] is None:
                            return Response({'error': 'Please fill in minimum bid step and maximum bid step.'})
                        if (
                            auctionType.max_price_settings is True
                            and auctionType.max_price_type == 2
                            and auctionType.individual_or_max_price == 2
                            and auctionItemData['max_price'] is None
                        ):
                            return Response({'error': 'Please fill value of max price for all items as chosing setting True'})

                        if auctionType.max_price_settings is False:
                            auctionItemData['max_price'] = None

                        if auctionType.individual_or_max_price == 1:
                            auctionItemData['max_price'] = None

                    if auction.auction_type1_id == 2 or auction.auction_type1_id == 5 or auction.auction_type1_id == 6:  # Sealed Bid #Dutch #Japanese
                        auctionItemData['price_yellow'] = None
                        auctionItemData['target_price'] = None
                        auctionItemData['minimum_bid_step'] = None
                        auctionItemData['maximum_bid_step'] = None
                        auctionItemData['max_price'] = None

                    # ---Case: Per position--------Calculate the average of each item and compare to max price of each item----------------
                    # ---Case: True, 2, 2 ------------------------------------------------------------------------
                    if (
                        auction.auction_type1_id == 1 or auction.auction_type1_id == 3 or auction.auction_type1_id == 4
                    ):  # add this line as Sealed bid dooes not have fields about max price
                        if auctionType.max_price_settings is True and auctionType.max_price_type == 2 and auctionType.individual_or_max_price == 2:
                            price_item = 0
                            for supplier in item['suppliers']:
                                price_item = price_item + supplier['price']

                            average_price_item = price_item / len(item['suppliers'])
                            bounded_max_price = 0.8 * average_price_item
                            error_max_and_start = abs(auctionItemData['max_price'] - average_price_item)

                            if error_max_and_start > bounded_max_price:
                                return Response(
                                    {
                                        'Note': 'You set max price per position.',
                                        'Error': 'Please fill value of max price not greater or lower than 80 percent average value of suppliers for each item.',
                                    }
                                )
                            else:
                                pass

                    auctionItemSerializer = AuctionItemSerializer(data=auctionItemData)

                    auctionItemSerializer.is_valid(raise_exception=True)
                    auctionItem = auctionItemSerializer.save()

                    # --------------------------------------------------------------------------

                    for supplier in item['suppliers']:

                        auctionSupplierData = {
                            **supplier,
                            'auction_id': auction.id,
                            'supplier_status': 1,
                        }

                        auctionSupplierSerializer = AuctionSupplierSerializer(data=auctionSupplierData)

                        auctionSupplierSerializer.is_valid(raise_exception=True)

                        if AuctionSupplier.objects.filter(auction_id=auction.id, user_id=supplier['user_id']).exists():
                            auctionSupplier = AuctionSupplier.objects.get(auction_id=auction.id, user_id=supplier['user_id'])
                        else:
                            auctionSupplier = auctionSupplierSerializer.save()

                        auctionItemSupplierData = {
                            **supplier,
                            'auction_id': auction.id,
                            'auction_item_id': auctionItem.id,
                            'auction_supplier_id': auctionSupplier.id,
                        }

                        auctionItemSupplierSerializer = AuctionItemSupplierSerializer(data=auctionItemSupplierData)

                        auctionItemSupplierSerializer.is_valid(raise_exception=True)
                        auctionItemSupplierSerializer.save()
                supplier_price_list = []
                for supplier in data['items'][0].get('suppliers'):
                    user_supplier = User.objects.select_related('supplier').get(id=supplier.get('user_id'))
                    if data.get('coupon'):
                        price = AuctionItemSupplier.objects.filter(auction=auction, auction_supplier__user=user_supplier).aggregate(Sum('price'))
                        supplier_price_list.append(price.get('price__sum'))

                    start_time = auction.start_time
                    tz = pytz.timezone(user_supplier.local_time)
                    start_time = start_time.replace(tzinfo=pytz.utc).astimezone(tz)

                    email_supplier_auction = EmailTemplates.objects.get(item_code="SupplierCreateAuction")
                    email_supplier = Template(email_supplier_auction.content).render(
                        Context(
                            {
                                "image": request.build_absolute_uri('/')[:-1] + "/static/logo_mail.png",
                                "name": user_supplier.full_name,
                                "company_long_name": user_buyer.buyer.company_full_name,
                                "auction_title": auction.title,
                                "auction_item_code": auction.item_code,
                                "start_time_date": start_time.strftime('%Y/%m/%d'),
                                "start_time": start_time.strftime('%H:%M'),
                                "AuctionID": auction.id,
                            }
                        )
                    )
                    try:
                        send_mail(
                            email_supplier_auction.title,
                            email_supplier,
                            "NextPro <no-reply@nextpro.io>",
                            [user_supplier.email],
                            html_message=email_supplier,
                            fail_silently=True,
                        )
                    except:
                        print("Fail mail")

                email_buyer_auction = EmailTemplates.objects.get(item_code="BuyerCreateAuction")
                title = Template(email_buyer_auction.title).render(
                    Context(
                        {
                            "auction_item_code": auction.item_code,
                        }
                    )
                )
                email_buyer = Template(email_buyer_auction.content).render(
                    Context(
                        {
                            "image": request.build_absolute_uri('/')[:-1] + "/static/logo_mail.png",
                            "name": user_buyer.full_name,
                            "auction_title": auction.title,
                            "auction_item_code": auction.item_code,
                        }
                    )
                )
                try:
                    send_mail(
                        title,
                        email_buyer,
                        "NextPro <no-reply@nextpro.io>",
                        [user_buyer.email],
                        html_message=email_buyer,
                        fail_silently=True,
                    )
                except:
                    print("Fail mail")

                # send coupon
                if data.get('coupon'):
                    coupon = Coupon.objects.filter(coupon_program=data.get('coupon'))
                    if coupon.exists():
                        coupon = coupon.first()
                        email_coupon = EmailTemplates.objects.get(item_code="CouponPublishedAuction")
                        title = Template(email_coupon.title).render(Context({'code': coupon.coupon_program}))
                        email = Template(email_coupon.content).render(
                            Context(
                                {
                                    "image": request.build_absolute_uri('/')[:-1] + "/static/logo_mail.png",
                                    "name": coupon.full_name,
                                    "buyer_name": user_buyer.full_name,
                                    "auction_title": auction.title,
                                    "auction_no": auction.item_code,
                                    "min_supplier_quote": str(round(min(supplier_price_list))) + " VND",
                                    "max_supplier_quote": str(round(max(supplier_price_list))) + " VND",
                                }
                            )
                        )
                        try:
                            send_mail(
                                title,
                                email,
                                "NextPro <no-reply@nextpro.io>",
                                [coupon.email],
                                html_message=email,
                                fail_silently=True,
                            )
                        except:
                            print("Fail mail")

                # send LimitAuction
                if (count + 1) == buyer_auction:
                    email_auction_limit = EmailTemplates.objects.get(item_code="LimitAuction")
                    email = Template(email_auction_limit.content).render(
                        Context(
                            {
                                "image": request.build_absolute_uri('/')[:-1] + "/static/logo_mail.png",
                                "name": user_buyer.full_name,
                                "link": "http://192.168.9.94:9001/account",
                            }
                        )
                    )
                    try:
                        send_mail(
                            email_auction_limit.title,
                            email,
                            "NextPro <no-reply@nextpro.io>",
                            [user_buyer.email],
                            html_message=email,
                            fail_silently=True,
                        )
                    except:
                        print("Fail mail")

                return Response({'success': True, 'auction_id': auction.id})
            except serializers.ValidationError as error:
                transaction.set_rollback(True)
                return Response({'error': error.detail, 'success': False}, status=error.status_code)

            except BaseException as error:
                transaction.set_rollback(True)
                print(error)
                return Response({'error': 'Server Error', 'success': False}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'error': 'Please upgrade your account to create eAuction!'})

class AuctionConfirm(APIView):
    def patch(self, request, pk):
        auction = Auction.objects.get(pk=pk)
        options = auction.get_options()

        auctionSupplier = AuctionSupplier.objects.get(auction_id=pk, user_id=request.user.id)
        auctionItem = AuctionItem.objects.filter(auction_id=pk)
        auctionItemsSupplier = AuctionItemSupplier.objects.filter(auction_id=pk, auction_supplier_id=auctionSupplier.id)
        for item in auctionItem:
            auctionItemSupplier = AuctionItemSupplier.objects.get(auction_id=pk, auction_item_id=item.id, auction_supplier_id=auctionSupplier.id)

        if auctionSupplier.supplier_status != 1:
            return Response({'success': False})

        if request.data['supplier_status'] not in (1, 3) or auctionSupplier.is_accept == True:
            return Response({'success': False})

        is_accept = True
        if request.data['supplier_status'] == 3:
            is_accept = False
        auctionSupplierSerializer = AuctionSupplierSerializer(
            auctionSupplier,
            data={
                'supplier_status': request.data['supplier_status'],
                'is_accept': is_accept,
                'note_from_supplier': request.data.get('note_from_supplier'),
                'is_accept_rule': request.data['is_accept'],
            },
            partial=True,
        )

        if request.data['supplier_status'] == 1:
            percent = 0.8
            # Check confirm price of each item that supplier type--------------------------

            if auction.auction_type1_id == 1 or auction.auction_type1_id == 3 or auction.auction_type1_id == 4:

                # Case 1: (False, Null, Null)
                if options.max_price_settings is False:
                    for data in request.data['items']:
                        for item in auctionItem:

                            auctionItemSupplier = AuctionItemSupplier.objects.get(
                                auction_id=pk, auction_item_id=item.id, auction_supplier_id=auctionSupplier.id
                            )

                            if auctionItemSupplier.auction_item_id == data['auction_item_id']:

                                if data['confirm_price'] is None:
                                    return Response({'error': 'Please fill value of changed starting price.'})

                                error_confirm_and_start = abs(data['confirm_price'] - auctionItemSupplier.price)
                                bounded_starting_price = auctionItemSupplier.price * percent

                                if error_confirm_and_start > bounded_starting_price:
                                    return Response({'error': 'Your new starting price not be greater or lower 80 percent than starting price.'})

                # Case 2: (True, 1, 1)
                if options.max_price_settings is True and options.max_price_type == 1 and options.individual_or_max_price == 1:

                    for data in request.data['items']:
                        for item in auctionItem:

                            auctionItemSupplier = AuctionItemSupplier.objects.get(
                                auction_id=pk, auction_item_id=item.id, auction_supplier_id=auctionSupplier.id
                            )

                            if auctionItemSupplier.auction_item_id == data['auction_item_id']:

                                if data['confirm_price'] is None:
                                    return Response({'error': 'Please fill value of changed starting price.'})

                    total_confirm_price = 0
                    for data in request.data['items']:
                        total_confirm_price = total_confirm_price + data['confirm_price']

                    total_starting_price = 0
                    for item_supplier in auctionItemsSupplier:
                        total_starting_price = total_starting_price + item_supplier.price

                    error_confirm_and_start = abs(total_confirm_price - total_starting_price)
                    bounded_confirm_price = total_starting_price * percent

                    if (total_confirm_price > total_starting_price) or (error_confirm_and_start > bounded_confirm_price):
                        return Response(
                            {
                                'error': 'Total of confirm price not be greater than total of your starting price and error not over 80 percent of starting price.'
                            }
                        )

                # Case 3: (True, 2, 1)
                if options.max_price_settings is True and options.max_price_type == 2 and options.individual_or_max_price == 1:

                    for data in request.data['items']:
                        for item in auctionItem:

                            auctionItemSupplier = AuctionItemSupplier.objects.get(
                                auction_id=pk, auction_item_id=item.id, auction_supplier_id=auctionSupplier.id
                            )

                            if auctionItemSupplier.auction_item_id == data['auction_item_id']:

                                if data['confirm_price'] is None:
                                    return Response({'error': 'Please fill value of changed starting price.'})

                                error_confirm_and_start = abs(data['confirm_price'] - auctionItemSupplier.price)
                                bounded_starting_price = auctionItemSupplier.price * percent

                                if (data['confirm_price'] > auctionItemSupplier.price) or (error_confirm_and_start > bounded_starting_price):
                                    return Response(
                                        {
                                            'error': 'Your new starting price not be greater than starting price and error not over 80 percent of starting price.'
                                        }
                                    )

                # Case 4: (True, 2, 2)
                if options.max_price_settings is True and options.max_price_type == 2 and options.individual_or_max_price == 2:
                    for data in request.data['items']:
                        for item in auctionItem:

                            auctionItemSupplier = AuctionItemSupplier.objects.get(
                                auction_id=pk, auction_item_id=item.id, auction_supplier_id=auctionSupplier.id
                            )
                            max_price_item = item.max_price

                            if auctionItemSupplier.auction_item_id == data['auction_item_id']:

                                if data['confirm_price'] is None:
                                    return Response({'error': 'Please fill value of changed starting price.'})

                                error_confirm_and_max = abs(data['confirm_price'] - max_price_item)
                                # percent = 0.8
                                bounded_max_price = max_price_item * percent

                                if (data['confirm_price'] > max_price_item) or (error_confirm_and_max > bounded_max_price):
                                    return Response(
                                        {
                                            'error': 'Your new starting price not be greater than max price and error not over 80 percent of maxw price.'
                                        }
                                    )

                # Case 5: (True, 1, 2)
                if options.max_price_settings is True and options.max_price_type == 1 and options.individual_or_max_price == 2:
                    for data in request.data['items']:
                        for item in auctionItem:

                            auctionItemSupplier = AuctionItemSupplier.objects.get(
                                auction_id=pk, auction_item_id=item.id, auction_supplier_id=auctionSupplier.id
                            )

                            if auctionItemSupplier.auction_item_id == data['auction_item_id']:

                                if data['confirm_price'] is None:
                                    return Response({'error': 'Please fill value of changed starting price.'})

                    total_confirm_price = 0
                    for data in request.data['items']:
                        total_confirm_price = total_confirm_price + data['confirm_price']

                    total_max_price = options.entire_auction

                    error_confirm_and_max = abs(total_confirm_price - total_max_price)
                    bounded_max_price = total_max_price * percent

                    if (total_confirm_price > total_max_price) or (error_confirm_and_max > bounded_max_price):
                        return Response(
                            {'error': 'Total of confirm price not be greater than max price and error not over 80 percent of starting price.'}
                        )

            if auction.auction_type1_id == 2:
                for data in request.data['items']:
                    for item in auctionItem:

                        auctionItemSupplier = AuctionItemSupplier.objects.get(
                            auction_id=pk, auction_item_id=item.id, auction_supplier_id=auctionSupplier.id
                        )

                        if auctionItemSupplier.auction_item_id == data['auction_item_id']:

                            if data['confirm_price'] is None:
                                return Response({'error': 'Please fill value of changed starting price.'})

                            error_confirm_and_start = abs(data['confirm_price'] - auctionItemSupplier.price)
                            bounded_starting_price = auctionItemSupplier.price * percent

                            if (error_confirm_and_start > bounded_starting_price) or (data['confirm_price'] > auctionItemSupplier.price):
                                return Response({'error': 'Your new starting price not be greater 80 percent than starting price.'})

        try:
            auctionSupplierSerializer.is_valid(raise_exception=True)
            auctionSupplierSerializer.save()
            amount = 0
            for data in request.data['items']:
                for item in auctionItem:
                    auctionItemSupplier = AuctionItemSupplier.objects.get(
                        auction_id=pk, auction_item_id=item.id, auction_supplier_id=auctionSupplier.id
                    )
                    if data['auction_item_id'] == item.id:

                        if request.data['supplier_status'] == 1:
                            auctionItemSupplierSerializer = AuctionItemSupplierSerializer(
                                auctionItemSupplier,
                                data={'auction_item_id': data['auction_item_id'], 'confirm_price': data['confirm_price']},
                                partial=True,
                            )
                            amount += float(data['confirm_price'])
                        if request.data['supplier_status'] == 3:
                            auctionItemSupplierSerializer = AuctionItemSupplierSerializer(
                                auctionItemSupplier, data={'auction_item_id': data['auction_item_id'], 'confirm_price': None}, partial=True
                            )
                        auctionItemSupplierSerializer.is_valid(raise_exception=True)
                        auctionItemSupplierSerializer.save()
            # payment
            if auction.auction_next_round is not None:
                payment_auction = PaymentAuction.objects.filter(
                    auction_id=auction.auction_next_round, history__user_payment__user=request.user
                ).first()
                payment_auction.auction = auction
                payment_auction.save()
                auctionSupplier.supplier_status = 5
                auctionSupplier.save()
                history = payment_auction.history
            else:
                platform_fee = PlatformFee.objects.all().first()
                amount_usd = amount
                auction_fee = AuctionFee.objects.filter(min_value__lte=amount_usd, max_value__gte=amount_usd)
                auction_fee = auction_fee.first()
                if auction_fee is None:
                    value = AuctionFee.objects.filter().aggregate(Max('max_value'))
                    auction_fee = AuctionFee.objects.filter(max_value=value.get('max_value__max')).first()
                amount = (amount * auction_fee.percentage) / 100
                amount_payment = amount + platform_fee.fee
                detail = [
                    {
                        'description': f'''Deposit {auction.item_code}''',
                        'quantity': 1,
                        'unit_price': amount_payment,
                        'total_amount': amount_payment,
                    }
                ]
                sub_total = round(amount_payment)
                charge_amount = round(platform_fee.fee * 1.1)
                amount_payment = round(amount_payment + amount_payment * 0.1)
                order_no = History.objects.filter().count() + 1
                order_no = '70' + str(order_no).zfill(4)
                history = History.objects.create(
                    user_payment=request.user.user_payment,
                    type=2,
                    transaction_description=f'''Deposit {auction.item_code}''',
                    balance=0,
                    status=1,
                    order_no=order_no,
                    amount=amount_payment,
                )
                payment_auction = PaymentAuction.objects.create(auction=auction, charge_amount=charge_amount, refund_amount=0, history=history)

                data = {
                    'user': request.user,
                    'invoice_no': history.order_no,
                    'invoice_date': datetime.strftime(history.date, '%d-%m-%Y'),
                    'detail_list': detail,
                    'sub_total': sub_total,
                    'is_topup': False,
                }

                invoice_generate(data)
                request_draft_invoice = f'''/{request.user.id}/{history.order_no}/draft_invoice.pdf'''
                history.request_draft_invoice = request_draft_invoice
                history.save()

            return Response({'success': True, 'auction_id': auction.id, 'history_id': history.id})

        except serializers.ValidationError as error:
            transaction.set_rollback(True)

            return Response({'error': error.detail, 'success': False}, status=error.status_code)

        except BaseException as error:
            transaction.set_rollback(True)
            print(error)
            return Response({'error': 'Server Error', 'success': False}, status=status.HTTP_400_BAD_REQUEST)


class AuctionFilter(FilterSet):
    status = filters.CharFilter(method="filter_by_status")
    item_code = filters.CharFilter('item_code')
    title = filters.CharFilter('title')
    id = filters.CharFilter('id')
    budget = filters.CharFilter('budget')
    start = filters.CharFilter(method='filter_by_start_time')
    end = filters.CharFilter(method='filter_by_end_time')
    currency = filters.CharFilter('currency')
    category = filters.CharFilter('category')
    username = filters.CharFilter(method='filter_by_username')
    full_name = filters.CharFilter(method='filter_by_full_name')
    company_full_name = filters.CharFilter(method='filter_by_company_full_name')

    class Meta:
        model = Auction
        fields = ['item_code', 'title', 'id', 'budget', 'start_time', 'end_time', 'status', 'currency', 'category']

    def filter_by_status(self, queryset, name, value):
        status_input = value.strip().split(',')
        queryset = queryset.filter(status__in=status_input)
        return queryset

    def filter_by_start_time(self, queryset, name, value):
        queryset = queryset.filter(start_time__lte=value)
        return queryset

    def filter_by_end_time(self, queryset, name, value):
        queryset = queryset.filter(end_time__gte=value)
        return queryset

    def filter_by_username(self, queryset, name, value):
        queryset = queryset.filter(user__username=value)
        return queryset

    def filter_by_full_name(self, queryset, name, value):
        queryset = queryset.filter(user__full_name=value)
        return queryset

    def filter_by_company_full_name(self, queryset, name, value):
        buyers = Buyer.objects.filter(company_full_name=value)
        list_id = []
        for buyer in buyers:
            list_id.append(buyer.user_id)
        queryset = queryset.filter(user_id__in=list_id)
        return queryset


class AuctionList(generics.ListAPIView):
    """
    Show the  E-Auction
    """

    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    ordering_fields = ('item_code', 'id', 'title', 'budget', 'start_time', 'end_time', 'status', 'currency', 'category')
    filterset_class = AuctionFilter

    def get_queryset(self):

        queryset = None
        if self.request.user.isAdmin():
            data = self.request.GET.get('ordering')
            if data is None:
                queryset = Auction.objects.filter().order_by('-id')
            elif data == 'username':
                queryset = Auction.objects.filter().order_by('user__username')
            elif data == '-username':
                queryset = Auction.objects.filter().order_by('-user__username')
            elif data == 'full_name':
                queryset = Auction.objects.filter().order_by('user__full_name')
            elif data == '-full_name':
                queryset = Auction.objects.filter().order_by('-user__full_name')
            elif data == 'company_full_name':
                queryset = Auction.objects.filter().order_by('user__buyer__company_full_name')
            elif data == '-company_full_name':
                queryset = Auction.objects.filter().order_by('-user__buyer__company_full_name')

        if self.request.user.isBuyer():
            queryset = Auction.objects.filter(user=self.request.user).order_by('-id')

        elif self.request.user.isSupplier():
            queryset = Auction.objects.filter(suppliers__user=self.request.user, status__gt=1).order_by('-id')
        return queryset

    def get_serializer_class(self):

        serializer_class = None

        if self.request.user.isAdmin():
            serializer_class = GetAuctionSerializer

        if self.request.user.isBuyer():
            serializer_class = GetAuctionSerializer

        elif self.request.user.isSupplier():
            serializer_class = AuctionBySupplierSerializer

        return serializer_class


class AuctionOptions(generics.RetrieveAPIView):

    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        auction = Auction.objects.get(pk=pk)

        if auction.auction_type1_id == 1:
            auctionType = AuctionTypeTrafficLight.objects.get(auction_id=auction.id)
            auctionTypeSerializer = AuctionTypeTrafficLightSerializer(auctionType)

            return Response(auctionTypeSerializer.data)

        if auction.auction_type1_id == 2:
            auctionType = AuctionTypeSealedBid.objects.get(auction_id=auction.id)
            auctionTypeSerializer = AuctionTypeSealedBidSerializer(auctionType)
            return Response(auctionTypeSerializer.data)

        if auction.auction_type1_id == 3:
            auctionType = AuctionTypeRanking.objects.get(auction_id=auction.id)
            auctionTypeSerializer = AuctionTypeRankingSerializer(auctionType)
            return Response(auctionTypeSerializer.data)

        if auction.auction_type1_id == 4:
            auctionType = AuctionTypePrices.objects.get(auction_id=auction.id)
            auctionTypeSerializer = AuctionTypePricesSerializer(auctionType)
            return Response(auctionTypeSerializer.data)

        if auction.auction_type1_id == 5:
            auctionType = AuctionTypeDutch.objects.get(auction_id=auction.id)
            auctionTypeSerializer = AuctionTypeDutchSerializer(auctionType)

            server_time = timezone.now()

            start_time = auction.start_time

            round_auction = auctionTypeSerializer.data['round_auction']
            price_validity = auctionTypeSerializer.data['price_validity']
            time_auction = round_auction * price_validity
            end_time = start_time + timedelta(seconds=time_auction)

            if server_time < start_time:
                current_round = None
            elif server_time == start_time:
                current_round = 1
            elif server_time > end_time:
                current_round = round_auction
            else:
                time_in_round = server_time - start_time
                time_in_round = time_in_round.total_seconds()
                current_round = time_in_round / price_validity
                current_round = math.ceil(current_round)

            auctionBid = AuctionBid.objects.filter(auction_id=auction.id).last()
            if auctionBid is None:
                last_bid_round = "N/A"
            else:
                time_last_bid = auctionBid.created
                time_in_round = time_last_bid - start_time
                time_in_round = time_in_round.total_seconds()
                last_bid_round = time_in_round / price_validity
                last_bid_round = math.ceil(last_bid_round)

            return Response(
                {
                    **auctionTypeSerializer.data,
                    'current_round': current_round,
                    'last_bid_round': last_bid_round,
                    'server_time': server_time,
                    'start_time': start_time,
                }
            )

        if auction.auction_type1_id == 6:
            auctionType = AuctionTypeJapanese.objects.get(auction_id=auction.id)
            auctionTypeSerializer = AuctionTypeJapaneseSerializer(auctionType)

            server_time = timezone.now()

            start_time = auction.start_time

            round_auction = auctionTypeSerializer.data['round_auction']
            price_validity = auctionTypeSerializer.data['price_validity']
            time_auction = round_auction * price_validity
            end_time = start_time + timedelta(seconds=time_auction)

            if server_time < start_time:
                current_round = None
            elif server_time == start_time:
                current_round = 1
            elif server_time > end_time:
                current_round = round_auction
            else:
                time_in_round = server_time - start_time
                time_in_round = time_in_round.total_seconds()
                current_round = time_in_round / price_validity
                current_round = math.ceil(current_round)

            auctionSuppliers = AuctionSupplier.objects.filter(auction_id=auction.id)

            suppliers = []
            for supplier in auctionSuppliers:
                user_id = supplier.user_id
                auctionBid = AuctionBid.objects.filter(auction_id=auction.id, user_id=user_id).last()

                if auctionBid is None:
                    last_bid_round = "N/A"
                else:
                    time_last_bid = auctionBid.created
                    time_in_round = time_last_bid - start_time
                    time_in_round = time_in_round.total_seconds()
                    last_bid_round = time_in_round / price_validity
                    last_bid_round = math.ceil(last_bid_round)

                suppliers.append({'user_id': user_id, 'last_bid_round': last_bid_round})

            return Response(
                {
                    **auctionTypeSerializer.data,
                    'current_round': current_round,
                    'suppliers': suppliers,
                    'server_time': server_time,
                    'start_time': start_time,
                }
            )


class AuctionDetail(generics.RetrieveUpdateDestroyAPIView):

    permission_classes = [permissions.IsAuthenticated]
    parsers_classes = [parsers.FormParser, parsers.MultiPartParser]

    def get_queryset(self):

        queryset = None

        if self.request.user.isAdmin():
            queryset = Auction.objects.filter().order_by('-id')

        if self.request.user.isBuyer():
            queryset = Auction.objects.filter(user=self.request.user).order_by('-id')

        elif self.request.user.isSupplier():
            queryset = Auction.objects.filter(suppliers__user=self.request.user).order_by('-id')

        return queryset

    def get_serializer_class(self):

        serializer_class = None

        if self.request.user.isAdmin():
            serializer_class = AuctionDetailSerializer

        if self.request.user.isBuyer():
            serializer_class = AuctionDetailSerializer

        elif self.request.user.isSupplier():
            serializer_class = AuctionDetailBySupplierSerializer

        return serializer_class

    def put(self, request, pk):
        auction = Auction.objects.get(pk=pk)
        try:
            auctionSerializer = AuctionSerializer(auction, data={**request.data, 'item_code': auction.item_code, 'user_id': request.user.id})
            auctionSerializer.is_valid(raise_exception=True)
            auctionSerializer.save()
            auction_items_mapping = AuctionItem.objects.filter(auction_id=auction.id)
            auctionTypeData = {**request.data.get('options'), 'auction_id': auction.id}

            # -------------check options---------------
            # Count suppliers of auction
            auction_items = request.data.get('items', [])
            number_of_items = len(auction_items)

            s = 0
            for item in auction_items:
                auction_suppliers = item['suppliers']
                number_of_suppliers = len(auction_suppliers)
                s = s + number_of_suppliers

            suppliers_of_auction = s / number_of_items

            # -------About-max(starting)-price--------------------------------------
            # -------Setting-max-price-of-Traffic-Light-or-Ranking-or-Prices-are-same-----------------
            if auction.auction_type1_id == 1 or auction.auction_type1_id == 3 or auction.auction_type1_id == 4:
                # --Case: False, Null, Null -----------------------------------
                if auctionTypeData['max_price_settings'] is False:
                    auctionTypeData['max_price_type'] = None
                    auctionTypeData['individual_or_max_price'] = None
                    auctionTypeData['entire_auction'] = None
                elif auctionTypeData['max_price_settings'] is True:
                    if auctionTypeData['max_price_type'] is None:
                        return Response({'error': 'Please choose type of max price when selecting True for setting max price.'})

                    if auctionTypeData['individual_or_max_price'] is None:
                        return Response({'error': 'Please check individual or maximum price'})

                    if auctionTypeData['max_price_type'] == 1 and auctionTypeData['individual_or_max_price'] == 2:
                        if auctionTypeData['entire_auction'] is None:
                            return Response({'error': 'Please fill max price for entire Auction.'})
                        else:
                            pass
                    else:
                        auctionTypeData['entire_auction'] = None

            # --------------------------------------------------------------------------------------------------

            if auction.auction_type1_id == 4:  # Prices
                # Check of bid mornitoring - views disabled and extension--------------------------------------

                if auctionTypeData['type_of_bid_monitoring6'] is False:
                    auctionTypeData['views_disabled'] = None
                    auctionTypeData['minutes'] = None
                elif auctionTypeData['type_of_bid_monitoring6'] is True:
                    if auctionTypeData['views_disabled'] == 1:
                        auctionTypeData['minutes'] = None
                        auctionTypeData['type_of_bid_monitoring1'] = None
                        auctionTypeData['type_of_bid_monitoring5'] = None

                    if auctionTypeData['views_disabled'] == 2 and auctionTypeData['minutes'] is None:
                        return Response({'error': 'Please fill the rest time you want to the information of suppliers be disabled.'})

                if auctionTypeData['auction_extension'] == 1:
                    auctionTypeData['auction_extension_trigger'] = None
                    auctionTypeData['number_of_rankings'] = None
                    auctionTypeData['frequency'] = None
                    auctionTypeData['trigger_time'] = None
                    auctionTypeData['prolongation_by'] = None
                else:
                    if auctionTypeData['type_of_bidding'] == 2:
                        auctionTypeData['auction_extension_trigger'] = 1
                        auctionTypeData['number_of_rankings'] = None
                        if (
                            auctionTypeData['frequency'] is None
                            or auctionTypeData['prolongation_by'] is None
                            or auctionTypeData['trigger_time'] is None
                        ):
                            return Response({'error': 'You must type the required field(s).'})
                    if auctionTypeData['type_of_bidding'] == 1:
                        if auctionTypeData['auction_extension_trigger'] is None:
                            return Response({'error': 'You must select the way how to trigger when chosing Automatic extension.'})

                        if auctionTypeData['auction_extension_trigger'] == 2:
                            if auctionTypeData['number_of_rankings'] is None:
                                return Response({'error': 'You must type the number of ranking when chosing the second way trigger.'})
                            if auctionTypeData['number_of_rankings'] > suppliers_of_auction:
                                return Response({'error': 'You must type a number lower than or equal to the number of suppliers.'})

                        if auctionTypeData['frequency'] is None or auctionTypeData['prolongation_by'] is None:
                            return Response({'error': 'You must type the required field(s).'})

                        if auctionTypeData['auction_extension_trigger'] != 2:
                            auctionTypeData['number_of_rankings'] = None

                        if auctionTypeData['auction_extension_trigger'] == 3:
                            auctionTypeData['trigger_time'] = None
                        else:
                            if auctionTypeData['trigger_time'] is None:
                                return Response({'error': 'You must type the required trigger time.'})

            # --------------------------------------------------------------------------------------------------

            if auction.auction_type1_id == 1 or auction.auction_type1_id == 3:  # Traffic Light or Ranking
                # ------Check-case-of-bid-mornitoring----------------------------------------
                # ------Check-3rd-then check-or-not-4th----------------------------------------
                # if auctionTypeData['type_of_bid_monitoring3'] is False

                # -------Check-6th-then-disabled-all-rest---------------------------------------
                if auctionTypeData['type_of_bid_monitoring6'] is False:
                    auctionTypeData['views_disabled'] = None
                    auctionTypeData['minutes'] = None
                elif auctionTypeData['type_of_bid_monitoring6'] is True:
                    if auctionTypeData['views_disabled'] == 1:
                        auctionTypeData['minutes'] = None
                        auctionTypeData['type_of_bid_monitoring1'] = None
                        auctionTypeData['type_of_bid_monitoring2'] = None
                        auctionTypeData['type_of_bid_monitoring3'] = None
                        auctionTypeData['type_of_bid_monitoring5'] = None

                    if auctionTypeData['views_disabled'] == 2 and auctionTypeData['minutes'] is None:
                        return Response({'error': 'Please fill the time you want to the information of suppliers be disabled.'})

                # -------------------------------------------------------------------------------------------------------------
                # ---------Type of bidding affect directly to extension trigger ------------------------------------------------

                if auctionTypeData['type_of_bid_monitoring3'] is False:
                    if auctionTypeData['type_of_bidding'] == 2:
                        return Response({'error': 'Default: Bidder sees the best bid '})
                if auctionTypeData['auction_extension'] == 1:
                    auctionTypeData['auction_extension_trigger'] = None
                    auctionTypeData['number_of_rankings'] = None
                    auctionTypeData['frequency'] = None
                    auctionTypeData['trigger_time'] = None
                    auctionTypeData['prolongation_by'] = None
                else:
                    if auctionTypeData['type_of_bidding'] == 2:
                        auctionTypeData['auction_extension_trigger'] = 1
                        auctionTypeData['number_of_rankings'] = None
                        if (
                            auctionTypeData['frequency'] is None
                            or auctionTypeData['prolongation_by'] is None
                            or auctionTypeData['trigger_time'] is None
                        ):
                            return Response({'error': 'You must type the required field(s).'})
                    if auctionTypeData['type_of_bidding'] == 1:
                        if auctionTypeData['auction_extension_trigger'] is None:
                            return Response({'error': 'You must select the way how to trigger when chosing Automatic extension.'})

                        if auctionTypeData['auction_extension_trigger'] == 2:
                            if auctionTypeData['number_of_rankings'] is None:
                                return Response({'error': 'You must type the number of ranking when chosing the second way trigger.'})
                            if auctionTypeData['number_of_rankings'] > suppliers_of_auction:
                                return Response({'error': 'You must type a number lower than or equal to the number of suppliers.'})

                        if auctionTypeData['frequency'] is None or auctionTypeData['prolongation_by'] is None:
                            return Response({'error': 'You must type the required field(s).'})

                        if auctionTypeData['auction_extension_trigger'] != 2:
                            auctionTypeData['number_of_rankings'] = None

                        if auctionTypeData['auction_extension_trigger'] == 3:
                            auctionTypeData['trigger_time'] = None
                        else:
                            if auctionTypeData['trigger_time'] is None:
                                return Response({'error': 'You must type the required trigger time.'})
            # ----------Check condition of Dutch Auction--------------------
            if auction.auction_type1_id == 5:
                if auctionTypeData['initial_price'] >= auctionTypeData['end_price']:
                    return Response({'error': 'Initial price must be lower than end price.'})

                if (
                    auctionTypeData['initial_price'] is None
                    or auctionTypeData['end_price'] is None
                    or auctionTypeData['price_step'] is None
                    or auctionTypeData['price_validity'] is None
                ):
                    return Response({'error': 'You must type the required field(s).'})

                if (auctionTypeData['initial_price'] + auctionTypeData['price_step']) > auctionTypeData['end_price']:
                    return Response({'error': 'You must type valid price step.'})

                round_auction = (auctionTypeData['end_price'] - auctionTypeData['initial_price']) / auctionTypeData['price_step']
                round_auction = math.ceil(round_auction)

            # ----------Check condition of Japanese Auction--------------------
            if auction.auction_type1_id == 6:
                if auctionTypeData['initial_price'] <= auctionTypeData['end_price']:
                    return Response({'error': 'Initial price must be greater than end price.'})

                if (
                    auctionTypeData['initial_price'] is None
                    or auctionTypeData['end_price'] is None
                    or auctionTypeData['price_step'] is None
                    or auctionTypeData['price_validity'] is None
                ):
                    return Response({'error': 'You must type the required field(s).'})

                if (auctionTypeData['initial_price'] - auctionTypeData['price_step']) < auctionTypeData['end_price']:
                    return Response({'error': 'You must type valid price step.'})

                round_auction = (auctionTypeData['initial_price'] - auctionTypeData['end_price']) / auctionTypeData['price_step']
                round_auction = math.ceil(round_auction)
            # --------------------------------------------------------------------------

            # delete type auction
            for number_type_id in range(1, 7):

                if number_type_id == 1 and number_type_id != auction.auction_type1_id:
                    auction_type1 = AuctionTypeTrafficLight.objects.filter(auction_id=auction.id).last()
                    if auction_type1 is not None:
                        auction_type1.delete()
                if number_type_id == 2 and number_type_id != auction.auction_type1_id:
                    auction_type2 = AuctionTypeSealedBid.objects.filter(auction_id=auction.id).last()
                    if auction_type2 is not None:
                        auction_type2.delete()
                if number_type_id == 3 and number_type_id != auction.auction_type1_id:
                    auction_type3 = AuctionTypeRanking.objects.filter(auction_id=auction.id).last()
                    if auction_type3 is not None:
                        auction_type3.delete()
                if number_type_id == 4 and number_type_id != auction.auction_type1_id:
                    auction_type4 = AuctionTypePrices.objects.filter(auction_id=auction.id).last()
                    if auction_type4 is not None:
                        auction_type4.delete()
                if number_type_id == 5 and number_type_id != auction.auction_type1_id:
                    auction_type5 = AuctionTypeDutch.objects.filter(auction_id=auction.id).last()
                    if auction_type5 is not None:
                        auction_type5.delete()
                if number_type_id == 6 and number_type_id != auction.auction_type1_id:
                    auction_type6 = AuctionTypeJapanese.objects.filter(auction_id=auction.id).last()
                    if auction_type6 is not None:
                        auction_type6.delete()

            # ------ Start to save E_Auction Type---------
            if auction.auction_type1_id == 1:  # Traffic Light
                auction_type = AuctionTypeTrafficLight.objects.filter(auction_id=auction.id).last()
                auctionTypeSerializer = AuctionTypeTrafficLightSerializer(auction_type, data=auctionTypeData)
            elif auction.auction_type1_id == 2:  # Sealed Bid
                auction_type = AuctionTypeSealedBid.objects.filter(auction_id=auction.id).last()
                auctionTypeSerializer = AuctionTypeSealedBidSerializer(auction_type, data=auctionTypeData)
            elif auction.auction_type1_id == 3:  # Ranking
                auction_type = AuctionTypeRanking.objects.filter(auction_id=auction.id).last()
                auctionTypeSerializer = AuctionTypeRankingSerializer(auction_type, data=auctionTypeData)
            elif auction.auction_type1_id == 4:  # Prices
                auction_type = AuctionTypePrices.objects.filter(auction_id=auction.id).last()
                auctionTypeSerializer = AuctionTypePricesSerializer(auction_type, data=auctionTypeData)
            elif auction.auction_type1_id == 5:
                auction_type = AuctionTypeDutch.objects.filter(auction_id=auction.id).last()
                auctionTypeSerializer = AuctionTypeDutchSerializer(auction_type, data={**auctionTypeData, 'round_auction': round_auction})
            elif auction.auction_type1_id == 6:
                auction_type = AuctionTypeJapanese.objects.filter(auction_id=auction.id).last()
                auctionTypeSerializer = AuctionTypeJapaneseSerializer(auction_type, data={**auctionTypeData, 'round_auction': round_auction})

            auctionTypeSerializer.is_valid(raise_exception=True)
            auctionType = auctionTypeSerializer.save()

            # ---Case: entire Auction--------Calculate the average of all items and compare to max price of all items----------------
            # ---Case: True, 1, 2---------------------database  updated-------------------------
            if (
                auction.auction_type1_id == 1 or auction.auction_type1_id == 3 or auction.auction_type1_id == 4
            ):  # add this line as Sealed bid dooes not have fields about max price
                if auctionType.max_price_settings is True and auctionType.max_price_type == 1 and auctionType.individual_or_max_price == 2:
                    average_price_total = 0
                    price_item = 0
                    for item in request.data['items']:
                        for supplier in item['suppliers']:
                            price_item = price_item + supplier['price']
                        average_price_total = price_item / len(item['suppliers'])
                    for item in data['items']:
                        item['max_price'] = None

                    max_price_all_items = auctionTypeData['entire_auction']

                    percent = 0.8
                    bounded_max_price = percent * average_price_total
                    error_max_and_average_total = abs(max_price_all_items - average_price_total)

                    if error_max_and_average_total > bounded_max_price:
                        return Response(
                            {
                                'Note': 'You choose max price for entire Auction.',
                                'error': 'Please fill in max price such that value of all items between 0,2 and 1,8 average of all items.',
                            }
                        )

            for item in request.data.get('items', []):
                auctionItemData = {**item, 'auction_id': auction.id}
                if auction.auction_type1_id == 1:  # Traffic Light
                    if (
                        auctionItemData['target_price'] is None
                        or auctionItemData['price_yellow'] is None
                        or auctionItemData['minimum_bid_step'] is None
                        or auctionItemData['maximum_bid_step'] is None
                    ):
                        return Response(
                            {
                                'error': 'Please check price level for yellow, target price, minimum bid step and maximum bid step, maybe one of these fields be incomplete'
                            }
                        )

                    if (
                        auctionType.max_price_settings is True
                        and auctionType.max_price_type == 2
                        and auctionType.individual_or_max_price == 2
                        and auctionItemData['max_price'] is None
                    ):
                        return Response({'error': 'Please fill value of max price for all items as chosing setting True'})

                    if auctionType.max_price_settings is False:
                        auctionItemData['max_price'] = None

                    if auctionType.individual_or_max_price == 1:
                        auctionItemData['max_price'] = None

                if auction.auction_type1_id == 3 or auction.auction_type1_id == 4:  # Ranking-Prices
                    auctionItemData['price_yellow'] = None

                    if auctionItemData['minimum_bid_step'] is None or auctionItemData['maximum_bid_step'] is None:
                        return Response({'error': 'Please fill in minimum bid step and maximum bid step.'})
                    if (
                        auctionType.max_price_settings is True
                        and auctionType.max_price_type == 2
                        and auctionType.individual_or_max_price == 2
                        and auctionItemData['max_price'] is None
                    ):
                        return Response({'error': 'Please fill value of max price for all items as chosing setting True'})

                    if auctionType.max_price_settings is False:
                        auctionItemData['max_price'] = None

                    if auctionType.individual_or_max_price == 1:
                        auctionItemData['max_price'] = None

                if auction.auction_type1_id == 2 or auction.auction_type1_id == 5 or auction.auction_type1_id == 6:  # Sealed Bid #Dutch #Japanese
                    auctionItemData['price_yellow'] = None
                    auctionItemData['target_price'] = None
                    auctionItemData['minimum_bid_step'] = None
                    auctionItemData['maximum_bid_step'] = None
                    auctionItemData['max_price'] = None

                # ---Case: Per position--------Calculate the average of each item and compare to max price of each item----------------
                # ---Case: True, 2, 2 ------------------------------------------------------------------------
                if (
                    auction.auction_type1_id == 1 or auction.auction_type1_id == 3 or auction.auction_type1_id == 4
                ):  # add this line as Sealed bid dooes not have fields about max price
                    if auctionType.max_price_settings is True and auctionType.max_price_type == 2 and auctionType.individual_or_max_price == 2:
                        price_item = 0
                        for supplier in item['suppliers']:
                            price_item = price_item + supplier['price']

                        average_price_item = price_item / len(item['suppliers'])
                        bounded_max_price = 0.8 * average_price_item
                        error_max_and_start = abs(auctionItemData['max_price'] - average_price_item)

                        if error_max_and_start > bounded_max_price:
                            return Response(
                                {
                                    'Note': 'You set max price per position.',
                                    'Error': 'Please fill value of max price not greater or lower than 80 percent average value of suppliers for each item.',
                                }
                            )
                        else:
                            pass
                # ----------update or add item ----------------------------------------------------------------

                item = None
                for auctionItem in auction_items_mapping:
                    if auctionItem.id == auctionItemData.get('id_items'):
                        item = auctionItem
                auctionItem = AuctionItemSerializer(item, data=auctionItemData)
                auctionItem.is_valid(raise_exception=True)
                auctionItemModel = auctionItem.save()

                # ----------update or add supplier----------------------------------------------------------------

                auction_suppliers_mapping = AuctionSupplier.objects.filter(auction_id=auction.id)

                for data_supplier in auctionItemData.get('suppliers', []):
                    supplier = None
                    for auction_supplier_mapping in auction_suppliers_mapping:
                        if auction_supplier_mapping.user_id == data_supplier.get('user_id'):
                            supplier = auction_supplier_mapping
                    auctionSupplier = AuctionSupplierSerializer(supplier, data={'auction_id': auction.id, **data_supplier})

                    auctionSupplier.is_valid(raise_exception=True)
                    auctionSupplierModel = auctionSupplier.save()

                    # ----------------- items suppliers-------------
                    auction_item_supplier = AuctionItemSupplier.objects.filter(
                        auction_supplier_id=auctionSupplierModel.id, auction_item_id=auctionItemModel.id
                    ).last()
                    auction_item_supplier1 = None
                    if auction_item_supplier is not None:
                        auction_item_supplier1 = auction_item_supplier
                    auctionItemSupplierSerializer = AuctionItemSupplierSerializer(
                        auction_item_supplier1,
                        data={
                            **data_supplier,
                            'auction_id': auction.id,
                            'auction_item_id': auctionItemModel.id,
                            'auction_supplier_id': auctionSupplierModel.id,
                        },
                    )
                    auctionItemSupplierSerializer.is_valid(raise_exception=True)
                    auctionItemSupplierSerializer.save()

            # --------------Delete items-------------
            auction_item = {data.id: data for data in auction_items_mapping}
            data_item = {data['id_items']: data for data in request.data.get('items', [])}
            for data in auction_item:
                if data not in data_item:

                    auction_item = AuctionItem.objects.get(pk=data, auction_id=auction.id)
                    auction_item_supplier = AuctionItemSupplier.objects.filter(auction_item_id=data)
                    auction_item_supplier.delete()
                    auction_item.delete()
            # -----------------------------Delete suppliers---------------------------------------------
            auction_suppliers_mapping = AuctionSupplier.objects.filter(auction_id=auction.id)
            auction_suppliers = {data.user_id: data for data in auction_suppliers_mapping}

            for x in request.data.get('items'):
                list_supplier = []
                for y in x.get('suppliers'):
                    list_supplier.append(y.get('user_id'))

            for user_id in auction_suppliers:
                if user_id not in list_supplier:
                    auction_supplier = AuctionSupplier.objects.get(auction_id=auction.id, user_id=user_id)
                    auction_item_supplier = AuctionItemSupplier.objects.filter(auction_supplier_id=auction_supplier.id)
                    auction_item_supplier.delete()
                    auction_supplier.delete()

            return Response({'success': True})
        except serializers.ValidationError as error:
            transaction.set_rollback(True)

            return Response({'error': error.detail, 'success': False}, status=error.status_code)


class AuctionResultsList(generics.RetrieveAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = None

        if self.request.user.user_type == 2:
            queryset = Auction.objects.filter(user=self.request.user)

        return queryset

    def get_serializer_class(self):
        serializer_class = None

        if self.request.user.user_type == 2:
            serializer_class = AuctionResultsSerializer

        return serializer_class

    def put(self, request, pk):
        auction = Auction.objects.get(pk=pk)

        if request.data.get('status') == 5 and request.data.get('reasons') is None:

            return Response({'error': 'You must type the reasons why this auction cancelled'})

        elif request.data.get('status') == 5 and request.data.get('reasons') is not None:
            suppliers = AuctionSupplier.objects.filter(auction_id=request.data.get('id'))

            for supplier in suppliers:
                supplier.awarded = 3
                supplier.save()

            auctionSerializer = AuctionResultsSerializer(auction, data=request.data)

            if auctionSerializer.is_valid():
                auctionSerializer.save()
            else:
                return Response(auctionSerializer.errors)
        else:
            auctionSerializer = AuctionResultsSerializer(auction, data=request.data)

            for data in request.data.get('suppliers'):
                if data.get('awarded') == 2 and data.get('reasons') is None:

                    return Response({'error': 'You must type the reasons why this supplier rejected'})
            if auctionSerializer.is_valid():
                auctionSerializer.save()

                suppliers = AuctionSupplier.objects.filter(auction_id=auctionSerializer.data.get('id'))

                supplier_mapping = {supplier.id: supplier for supplier in suppliers}
                data_mapping = {data['id']: data for data in request.data.get('suppliers')}

                for id, data in data_mapping.items():
                    supplier = supplier_mapping.get(id)

                    supplierSerializer = AuctionSupplierResultSerializer(supplier, data=data)

                    if supplierSerializer.is_valid():
                        supplierSerializer.save()
                    else:
                        return Response(supplierSerializer.errors)
            else:
                return Response(auctionSerializer.errors)

        return Response(auctionSerializer.data)


class CurrentTimestamp(generics.RetrieveAPIView):
    def get(self, format=None):

        server_time = datetime.strftime(timezone.now(), '%Y-%m-%dT%H:%M:%SZ')

        return Response(
            {
                'server_time': server_time,
            }
        )


def get_next_round(auction, list_auction):
    auction = Auction.objects.filter(auction_next_round=auction).first()
    if auction:
        list_auction.append(auction)
        return get_next_round(auction, list_auction)
    return list_auction


class AuctionReportExportCoupon(APIView):
    def get(self, request, format=None):
        auctions = Auction.objects.filter(auction_next_round__isnull=True)
        list_auction = []
        for auction in auctions:
            list_auction.append(auction)
            list_auction = get_next_round(auction, list_auction)

        serializer = AuctionReportExportSerializer(list_auction, many=True)
        excel_data = [
            [
                'eAuction Code',
                'Buyer Code',
                'Buyer Name',
                'Purchasing Organization',
                'E-Auction Title',
                'Category',
                'Country',
                'City',
                'Value before OBE',
                'Best bid after OBE',
                'Budget',
                'Currency',
                'Auction type',
                'Test/Live',
                'No. of extension round',
                'Start',
                'End',
                'Awarded date',
                'Awarded supplier',
                'Status',
                'Invited',
                'Participant',
                'Rejected',
                'Award order',
                'Saving amount vs. budget',
                'Saving amount vs. budget (%)',
                'eAuction Fee',
                'Coupon Code',
                'Description',
                'Commission %',
                'Commission fee',
                'Valid From',
                'Valid To',
                'Note',
                'Ambassadors Email',
                'Status ',
            ],
        ]
        for auction in serializer.data:
            excel_data.append(
                [
                    auction.get('item_code'),
                    auction.get('username'),
                    auction.get('full_name'),
                    auction.get('company_full_name'),
                    auction.get('title'),
                    auction.get('category'),
                    str(auction.get('country')),
                    auction.get('city'),
                    auction.get('value_before_OBE'),
                    auction.get('best_bid_after_OBE'),
                    "{:10,.2f}".format(auction.get('budget')),
                    auction.get('currency'),
                    str(auction.get('auction_type')),
                    auction.get(''),  #'Test/Live'
                    auction.get('number_next_round'),
                    auction.get('start_time'),
                    auction.get('end_time'),
                    auction.get('award_date'),
                    auction.get('award_supplier'),
                    auction.get('status'),
                    auction.get('invited'),
                    auction.get('participant'),
                    auction.get('rejected'),
                    auction.get('award_order'),
                    auction.get('amount_vs_budget'),
                    auction.get('amount_vs_budget_percent'),
                    auction.get('auction_fee'),
                    auction.get('coupon_code'),
                    auction.get('coupon_description'),
                    auction.get('coupon_commission'),
                    auction.get('coupon_commission_fee'),
                    auction.get('coupon_valid_from'),
                    auction.get('coupon_valid_to'),
                    auction.get('coupon_note'),
                    auction.get('coupon_email'),
                    auction.get('coupon_status'),
                ]
            )

        wb = Workbook()
        for row in excel_data:
            wb.active.append(row)
        font = Font(size=13, bold=True)
        fill = PatternFill(bgColor="ffffcc")
        ws = wb.active
        for cell in ws["1:1"]:
            cell.font = font
            cell.fill = fill

        for column_cells in ws.columns:
            unmerged_cells = list(filter(lambda cell_to_check: cell_to_check.coordinate not in ws.merged_cells, column_cells))
            length = max(len(str(cell.value)) for cell in unmerged_cells)
            ws.column_dimensions[unmerged_cells[0].column_letter].width = length * 1.5
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=CouponData.xlsx'
        wb.save(response)
        return response


class AuctionReports:
    def report(auction, language_code):
        wb = Workbook()
        sheet1(wb=wb, auction=auction, language_code=language_code)
        sheet2(wb=wb, auction=auction, language_code=language_code)
        sheet3(wb=wb, auction=auction, language_code=language_code)
        wb.remove(wb.get_sheet_by_name('Sheet'))
        Path(os.path.join(settings.MEDIA_ROOT, 'auction_reports')).mkdir(parents=True, exist_ok=True)
        file_path = os.path.join(settings.MEDIA_ROOT, 'auction_reports', f'''{auction.item_code}.xlsx''')

        wb.save(file_path)


def sheet1(wb, auction, language_code):
    label = {
        "vi": {
            "Description_type": "STT/Mô tả",
            "Header": "Tiêu đề",
            "Buyer": "Bên mua:",
            "Running_Time": "Thời gian diễn ra:",
            "Status": "Tình trạng:",
            "Currency": "Đơn vị tính:",
            "Invited": "Đã mời:",
            "Participant": "Tham dự:",
            "Rejected": "Từ chối:",
            "Offer_Submission": "Nộp hồ sơ:",
            "until": " đến ",
            "Commercial_Term": "Điều khoản thương mại",
            "Contract_Type": "Loại hợp đồng:",
            "Payment_Term": "Điều khoản thanh toán:",
            "Delivery_Term": "Điều khoản thanh toán:",
            "Contract_Features": "Tính năng hợp đồng:",
            "Awarding_Mode": "Quy chế trao thầu",
            "Mode": "Quy chế",
            "Auction_Mode": "Quy chế đấu giá:",
            "Starting_Prices": "Giá khởi điểm:",
            "Type_of_bidding": "Loại dự thầu:",
            "Extension": "Gia hạn:",
            "Trigger_Time": "Thời gian kích hoạt:",
            "Prolongation_By": "Kéo dài:",
            "Extension_trigger": "Gia hạn kích hoạt:",
            "max": "Tối đa ",
            "min": " phút",
        },
        "en": {
            "Description_type": "No./Description of type",
            "Header": "Header",
            "Buyer": "Buyer:",
            "Running_Time": "Running time:",
            "Status": "Status:",
            "Currency": "Calculation currency:",
            "Invited": "Invited:",
            "Participant": "Participant",
            "Rejected": "Rejected",
            "Offer_Submission": "Offer_Submission",
            "until": " until ",
            "Commercial_Term": "Commercial Term",
            "Contract_Type": "Contract Type",
            "Payment_Term": "Payment Term:",
            "Delivery_Term": "Delivery Term:",
            "Contract_Features": "Contract Features:",
            "Awarding_Mode": "Awarding Mode",
            "Mode": "Mode",
            "Auction_Mode": "Auction Mode:",
            "Starting_Prices": "Starting Prices:",
            "Type_of_bidding": "Type of bidding:",
            "Extension": "Extension:",
            "Trigger_Time": "Trigger Time:",
            "Prolongation_By": "Prolongation By:",
            "Extension_trigger": "Extension trigger:",
            "max": "Max.",
            "min": " min",
        },
    }

    tz = pytz.timezone(auction.user.local_time)
    start_time = auction.start_time.replace(tzinfo=pytz.utc).astimezone(tz)
    start_time = datetime.strftime(start_time, '%d-%b-%Y %H:%M:%S')
    end_time = auction.extend_time or auction.end_time
    end_time = end_time.replace(tzinfo=pytz.utc).astimezone(tz)
    end_time = datetime.strftime(end_time, '%d-%b-%Y %H:%M:%S')

    contract_type = auction.contract_type
    if ContractTypeTranslation.objects.filter(contract_type=contract_type, language_code=language_code).exists():
        contract_type = ContractTypeTranslation.objects.filter(contract_type=contract_type, language_code=language_code).first()
    name = contract_type.name
    if auction.contract_type_id == 2:
        contract_type = name + " " + datetime.strftime(auction.contract_from, '%d.%m.%Y') + " - " + datetime.strftime(auction.contract_to, '%d.%m.%Y')
    else:
        contract_type = name

    if auction.status == 4 and language_code == 'en':
        status = 'Finished'
    elif auction.status == 4 and language_code == 'vi':
        status = 'Đã kết thúc'
    else:
        status = None
    payment_term = auction.payment_term
    if PaymentTermTranslation.objects.filter(payment_term=payment_term, language_code=language_code).exists():
        payment_term = PaymentTermTranslation.objects.filter(payment_term=payment_term, language_code=language_code).first()
    payment_term = payment_term.name
    delivery_term = auction.delivery_term
    if DeliveryTermTranslation.objects.filter(delivery_term=delivery_term, language_code=language_code).exists():
        delivery_term = DeliveryTermTranslation.objects.filter(delivery_term=delivery_term, language_code=language_code).first()
    delivery_term = delivery_term.name
    technical_weighting = auction.technical_weighting
    if TechnicalWeightingTranslation.objects.filter(technical_weighting=technical_weighting, language_code=language_code).exists():
        technical_weighting = TechnicalWeightingTranslation.objects.filter(
            technical_weighting=technical_weighting, language_code=language_code
        ).first()
    technical_weighting = technical_weighting.name
    if auction.auction_type1_id in (1, 3, 4):
        individual_or_max_price = auction.get_options().individual_or_max_price
        type_of_bibbing = auction.get_options().type_of_bidding
        if individual_or_max_price == 1:
            individual_or_max_price = "Individual max prices per supplier"
        elif individual_or_max_price == 2:
            individual_or_max_price = "Maximum price must not be reached"
        else:
            individual_or_max_price = None

        if type_of_bibbing == 1:
            type_of_bibbing = "Bidder must underbid himself"
        elif type_of_bibbing == 2:
            type_of_bibbing = "Bidder must underbid the best bid"
        else:
            type_of_bibbing = None
        prolongation_by = auction.get_options().prolongation_by
        trigger_time = auction.get_options().trigger_time
        frequency = auction.get_options().frequency
        auction_extension_trigger = auction.get_options().auction_extension_trigger
        if auction_extension_trigger == 1:
            auction_extension_trigger = "New bid in trigger time"
        elif auction_extension_trigger == 2:
            auction_extension_trigger = "New bid among the best rankings"
        elif auction_extension_trigger == 3:
            auction_extension_trigger = "Equal new best bids"
        else:
            auction_extension_trigger = None
    else:
        individual_or_max_price = None
        type_of_bibbing = None
        prolongation_by = None
        trigger_time = None
        frequency = None
        auction_extension_trigger = None
    auction.currency.name
    currency = auction.currency
    if CurrencyTranslation.objects.filter(currency=currency, language_code=language_code).exists():
        currency = CurrencyTranslation.objects.filter(currency=currency, language_code=language_code).first()
    currency = currency.name

    ws1 = wb.create_sheet("General Information", 0)
    ws1.sheet_view.showGridLines = False
    ws1.merge_cells("A2:C3")

    ws1.cell(row=2, column=1, value=label.get(language_code).get('Description_type')).alignment = Alignment(vertical="center")
    ws1['D2'] = f'''{auction.item_code} - {auction.title} '''
    ws1['D3'] = str(auction.auction_type1)

    img = Image(os.path.join(settings.STATICFILES_DIRS[0], 'logo_report.png'))

    img.height = 40
    img.width = 300

    ws1.merge_cells("F2:I3")
    ws1.add_image(img, 'F2')
    ws1.column_dimensions["E"].width = 32.0
    set_background_black(ws=ws1, min_row=2, max_col=10, max_row=3)

    # header
    set_border(ws1, "A6:I16")
    ws1["A5"] = label.get(language_code).get('Header')
    set_background_black(ws=ws1, min_row=5, max_col=3, max_row=5)
    ws1.merge_cells("A5:C5")
    ws1.cell(row=7, column=1, value=label.get(language_code).get('Buyer')).font = Font(bold=True)
    ws1.cell(row=9, column=1, value=label.get(language_code).get('Running_Time')).font = Font(bold=True)
    ws1.cell(row=10, column=1, value=label.get(language_code).get('Status')).font = Font(bold=True)
    ws1.cell(row=11, column=1, value=label.get(language_code).get('Currency')).font = Font(bold=True)
    ws1.cell(row=12, column=1, value=label.get(language_code).get('Invited')).font = Font(bold=True)
    ws1.cell(row=13, column=1, value=label.get(language_code).get('Participant')).font = Font(bold=True)
    ws1.cell(row=14, column=1, value=label.get(language_code).get('Rejected')).font = Font(bold=True)
    ws1.cell(row=15, column=1, value=label.get(language_code).get('Offer_Submission')).font = Font(bold=True)
    ws1.cell(row=7, column=4, value=auction.user.buyer.company_full_name).font = Font(color="FF0000")
    ws1.cell(row=8, column=4, value=auction.user.full_name).font = Font(color="FF0000")
    ws1['D9'] = str(start_time) + label.get(language_code).get('until') + str(end_time)
    ws1['D10'] = status
    ws1['D11'] = currency
    ws1['D12'] = AuctionSupplier.objects.filter(auction=auction).count()
    ws1['D13'] = AuctionSupplier.objects.filter(auction=auction, is_accept=True).count()
    ws1['D14'] = AuctionSupplier.objects.filter(auction=auction, is_accept=False).count()
    ws1['D15'] = AuctionBid.objects.filter(auction_id=auction.id).count()

    # Commercial Term
    set_border(ws1, "A19:I23")
    ws1["A18"] = label.get(language_code).get('Commercial_Term')
    set_background_black(ws=ws1, min_row=18, max_col=3, max_row=18)
    ws1.merge_cells("A18:C18")
    ws1.cell(row=19, column=1, value=label.get(language_code).get('Contract_Type')).font = Font(bold=True)
    ws1.cell(row=20, column=1, value=label.get(language_code).get('Payment_Term')).font = Font(bold=True)
    ws1.cell(row=21, column=1, value=label.get(language_code).get('Delivery_Term')).font = Font(bold=True)
    ws1.cell(row=22, column=1, value=label.get(language_code).get('Contract_Features')).font = Font(bold=True)

    ws1["D19"] = contract_type
    ws1["D20"] = payment_term
    ws1["D21"] = delivery_term
    ws1["D22"] = None

    # Awarding Mode
    set_border(ws1, "A26:I27")
    ws1["A25"] = label.get(language_code).get('Awarding_Mode')
    set_background_black(ws=ws1, min_row=25, max_col=3, max_row=25)
    ws1.merge_cells("A25:C25")
    ws1.cell(row=26, column=1, value=label.get(language_code).get('Auction_Mode')).font = Font(bold=True)
    ws1["D26"] = technical_weighting

    # Auction mode
    set_border(ws1, "A30:I36")
    ws1["A29"] = label.get(language_code).get('Auction_Mode')
    set_background_black(ws=ws1, min_row=29, max_col=3, max_row=29)
    ws1.merge_cells("A29:C29")
    ws1.cell(row=30, column=1, value=label.get(language_code).get('Starting_Prices')).font = Font(bold=True)
    ws1.cell(row=31, column=1, value=label.get(language_code).get('Type_of_bidding')).font = Font(bold=True)
    ws1.cell(row=32, column=1, value=label.get(language_code).get('Extension')).font = Font(bold=True)
    ws1.cell(row=33, column=1, value=label.get(language_code).get('Trigger_Time')).font = Font(bold=True)
    ws1.cell(row=34, column=1, value=label.get(language_code).get('Prolongation_By')).font = Font(bold=True)
    ws1.cell(row=35, column=1, value=label.get(language_code).get('Extension_trigger')).font = Font(bold=True)

    ws1.cell(row=30, column=4, value=individual_or_max_price).font = Font(color="00CC00")
    ws1.cell(row=31, column=4, value=type_of_bibbing).font = Font(color="FF0000")
    if frequency:
        frequency = label.get(language_code).get('max') + str(frequency)
    if trigger_time:
        trigger_time = str(trigger_time) + label.get(language_code).get('min')
    if prolongation_by:
        prolongation_by = str(prolongation_by) + label.get(language_code).get('min')
    ws1["D32"] = frequency
    ws1["D33"] = trigger_time
    ws1["D34"] = prolongation_by
    ws1["D35"] = auction_extension_trigger
    for rows in ws1.iter_rows(min_row=32, max_row=35, min_col=4, max_col=9):
        for cell in rows:
            cell.font = Font(color="000000")
            cell.fill = PatternFill(fgColor="FFFF00", fill_type="solid")


def sheet2(wb, auction, language_code):
    auction_item = AuctionItem.objects.filter(auction=auction)
    auction_supplier = AuctionSupplier.objects.filter(auction=auction, is_accept=True, supplier_status__in=[5, 6, 8, 9, 10]).order_by('id')

    supplier_sort = []
    auction_result = {}
    count_reach_target = 0
    list_total_bid = []
    list_total_confirm_price_supplier = []
    for supplier in auction_supplier:
        is_accept_rule = supplier.is_accept_rule
        total_bid = 0
        total_confirm_price = 0
        result = {"user": supplier.user, "auction_supplier": supplier}
        sub_count_reach_target = 0
        total_target = 0
        technical_score_supplier = 0
        for item in auction_item:
            auction_bids = AuctionBid.objects.filter(user=supplier.user, auction_item=item).order_by('id')
            auction_item_supplier = AuctionItemSupplier.objects.filter(auction_item=item, auction_supplier=supplier).first()
            if auction_bids.exists():
                price = auction_bids.last().price
            else:
                price = auction_item_supplier.confirm_price
            total_bid += price
            total_confirm_price += auction_item_supplier.confirm_price
            if auction_item_supplier.technical_score:
                technical_score_supplier += auction_item_supplier.technical_score
            #  check bid with target
            if item.target_price:
                total_target += item.target_price
                if price <= item.target_price:
                    sub_count_reach_target += 1
        if len(auction_item) == sub_count_reach_target:
            count_reach_target += 1
        list_total_bid.append(total_bid)
        result['technical_score_supplier'] = technical_score_supplier
        result['total_bid'] = total_bid
        result['total_confirm_price'] = total_confirm_price
        result['is_accept_rule'] = is_accept_rule
        auction_result["total_target"] = total_target
        list_total_confirm_price_supplier.append(total_confirm_price)
        supplier_sort.append(result)
    auction_result["count_reach_target"] = count_reach_target
    auction_result["list_total_confirm_price_supplier"] = list_total_confirm_price_supplier
    items = []
    total_reach_target = 0
    total_budget = 0
    for item in auction_item:
        total_budget += item.budget
        auction_bid = {}
        auction_bid['item'] = item
        auction_bid['suppliers'] = []
        total_reach_target_list = []
        for supplier in supplier_sort:
            auction_bids = AuctionBid.objects.filter(user=supplier.get('user'), auction_item=item).order_by('id')
            auction_item_supplier = AuctionItemSupplier.objects.filter(auction_item=item, auction_supplier=supplier.get('auction_supplier')).first()
            if auction_bids.exists():
                price = auction_bids.last().price
            else:
                price = auction_item_supplier.confirm_price
            auction_bid['suppliers'].append(
                {
                    'user_id': supplier.get('user'),
                    'price': price,
                    'confirm_price': auction_item_supplier.confirm_price,
                }
            )
            total_reach_target_list.append(price)
        items.append(auction_bid)
        if len(total_reach_target_list) != 0:
            total_reach_target += min(total_reach_target_list)
    auction_result["total_budget"] = total_budget
    auction_result["total_reach_target"] = total_reach_target
    if auction.split_order == 1 and len(supplier_sort) > 0:
        auction_result["best_bid"] = min(list_total_bid)
    elif auction.split_order == 2 and len(supplier_sort) > 0:
        auction_result["best_bid"] = total_reach_target
    else:
        auction_result["best_bid"] = None

    ws2 = wb.create_sheet("Sourcing Recommendation", 1)
    ws2.sheet_view.showGridLines = False
    label = {
        "vi": {
            "Description_type": "STT/Mô tả",
            "Auction_Result": "Kết quả đấu thầu	",
            "Best_Bid": "Dự thầu tốt nhất		",
            "Saving_Budget_Volume": "Tiết kiệm so với Khối lượng Ngân sách",
            "Savings_best_1st_offers": "Tiết kiệm so với ưu đãi tốt nhất đầu tiên",
            "Reach_Target": "Mục tiêu đạt được",
            "Reach_Target_%": "Mục tiêu đạt được %	",
            "Sourcing_Decision": "Quyết định tìm nguồn cung ứng",
            "Qualified_supplier_Name": "Tên nhà cung cấp đủ điều kiện",
            "Acceptance_of_auction_rules": "Chấp nhận các quy tắc đấu giá",
            "Item": "Mục",
            "Item_name": "Tên mặt hàng	",
            "Quantity": "Số lượng",
            "Currency": "Đơn vị tiền",
            "Starting_Price": "Giá khởi điểm",
            "Auction_Price": "Giá thầu",
            "Price_per_PU": "Giá mỗi PU",
            "Total_price": "Tổng giá",
            "Total_Amount_for_Total_Quantity": "Tổng số tiền cho Tổng số lượng:",
            "Sourcing_Recommendation": "Khuyến nghị tìm nguồn cung ứng:",
            "Recommendation_Sourcing_Decision": "Đề xuất Quyết định Tìm nguồn cung ứng:",
            "Others_Comments": "Các bình luận khác",
            "Accepted": "Được chấp nhận	",
            "Bidders": "  Nhà thầu",
            "First": "Nhất",
            "Second": "Nhì",
            "Third": "Ba",
            "Award": "Trao thầu",
            "Total_technical_point_100": "Tổng số điểm kỹ thuật/100: ",
            "Commercial_Against_Technical": "Giá trị của điểm thương mại:",
            "Bonus_Malus_value": "Giá trị điểm thưởng / điểm trừ:",
            "Total_Amount_for_Total_Quantity_After_Bonus": "Tổng giá trị sau khi trừ điểm thưởng/điểm trừ:",
            "Budget": "Ngân sách",
        },
        "en": {
            "Description_type": "No./Description of type",
            "Auction_Result": "Auction Result",
            "Best_Bid": "Best Bid",
            "Saving_Budget_Volume": "Savings vs. Budget Volume",
            "Savings_best_1st_offers": "Savings vs. The best 1st offers:",
            "Reach_Target": "Reach Target",
            "Reach_Target_%": "Reach Target %",
            "Sourcing_Decision": "Sourcing Decision",
            "Qualified_supplier_Name": "Qualified supplier Name",
            "Acceptance_of_auction_rules": "Acceptance of auction rules ",
            "Item": "Item",
            "Item_name": "Item name",
            "Quantity": "Quantity:",
            "Currency": "Currency ",
            "Starting_Price": "Starting Price",
            "Auction_Price": "Auction Price",
            "Price_per_PU": "Price per PU",
            "Total_price": "Total price",
            "Total_Amount_for_Total_Quantity": "Total Amount for Total Quantity",
            "Sourcing_Recommendation": "Sourcing Recommendation:",
            "Recommendation_Sourcing_Decision": "Recommendation Sourcing Decision",
            "Others_Comments": "Others Comments:",
            "Accepted": "Accepted",
            "Bidders": " Bidders",
            "First": "First",
            "Second": "Second",
            "Third": "Third",
            "Award": "Award",
            "Total_technical_point_100": 'Total technical point/100:',
            "Commercial_Against_Technical": 'Commercial Against Technical:',
            "Bonus_Malus_value": "Bonus/Malus value:",
            "Total_Amount_for_Total_Quantity_After_Bonus": "Total Amount for Total Quantity After Bonus:",
            "Budget": "Budget",
        },
    }

    ws2.merge_cells("A2:C3")
    ws2.cell(row=2, column=1, value=label.get(language_code).get('Description_type')).alignment = Alignment(vertical="center")
    ws2['D2'] = f'''{auction.item_code} - {auction.title} '''
    ws2['D3'] = str(auction.auction_type1)

    img = Image(os.path.join(settings.STATICFILES_DIRS[0], 'logo_report.png'))

    img.height = 40
    img.width = 300
    ws2.column_dimensions["D"].width = 20.0

    ws2.merge_cells("F2:I3")
    ws2.add_image(img, 'N2')
    ws2.column_dimensions["C"].width = 15.0
    set_background_black(ws=ws2, min_row=2, max_col=18, max_row=3)

    # Auction Result
    auction_results(ws2=ws2, auction_result=auction_result, auction=auction, label=label, language_code=language_code)

    # Sourcing Decision
    set_border(ws2, "A10:C10")
    ws2["A10"] = label.get(language_code).get('Sourcing_Decision')
    set_background_black(ws=ws2, min_row=10, max_col=3, max_row=10)
    ws2.merge_cells("A10:C10")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws2['A11'] = label.get(language_code).get('Qualified_supplier_Name')

    ws2.merge_cells("A11:G11")
    ws2['A12'] = label.get(language_code).get('Acceptance_of_auction_rules')
    ws2.merge_cells("A12:G12")

    ws2.cell(row=13, column=1, value=label.get(language_code).get('Item')).alignment = Alignment(vertical="center")
    ws2.cell(row=13, column=1).font = Font(bold=True)
    ws2.merge_cells("A13:A14")
    ws2.cell(row=13, column=2, value=label.get(language_code).get('Item_name')).alignment = Alignment(vertical="center")
    ws2.cell(row=13, column=2).font = Font(bold=True)
    ws2.merge_cells("B13:C14")
    ws2.cell(row=13, column=4, value="PU").alignment = Alignment(vertical="center")
    ws2.cell(row=13, column=4).font = Font(bold=True)
    ws2.merge_cells("D13:D14")
    ws2.cell(row=13, column=5, value=label.get(language_code).get('Quantity')).alignment = Alignment(vertical="center")
    ws2.cell(row=13, column=5).font = Font(bold=True)
    ws2.merge_cells("E13:E14")
    ws2.cell(row=13, column=6, value=label.get(language_code).get('Currency')).alignment = Alignment(vertical="center")
    ws2.cell(row=13, column=6).font = Font(bold=True)

    ws2.merge_cells("F13:F14")
    ws2.cell(row=13, column=7, value=label.get(language_code).get('Budget')).alignment = Alignment(vertical="center")
    ws2.cell(row=13, column=7).font = Font(bold=True)
    ws2.merge_cells("G13:G14")

    start_column_11 = 8
    end_column_11 = 11
    start_column_12 = 8
    end_column_12 = 9
    for supplier in supplier_sort:
        ws2.merge_cells(start_row=11, start_column=start_column_11, end_row=11, end_column=end_column_11)
        ws2.cell(row=11, column=start_column_11).font = Font(bold=True)
        ws2.cell(row=11, column=start_column_11, value=supplier.get('user').full_name).alignment = Alignment(horizontal="center")
        if supplier.get('is_accept_rule'):
            is_accept_rule = label.get(language_code).get('Accepted')
        else:
            is_accept_rule = None
        ws2.cell(row=12, column=start_column_11 + 2, value=is_accept_rule).font = Font(color="32a85e")

        for i in range(2):
            ws2.merge_cells(start_row=12, start_column=start_column_12, end_row=12, end_column=end_column_12)
            ws2.merge_cells(start_row=13, start_column=start_column_12, end_row=13, end_column=end_column_12)
            if i == 0:
                ws2.cell(row=13, column=start_column_12, value=label.get(language_code).get('Starting_Price')).alignment = Alignment(
                    horizontal="center"
                )
                ws2.cell(row=13, column=start_column_12).font = Font(bold=True)

            if i == 1:
                ws2.cell(row=13, column=start_column_12, value=label.get(language_code).get('Auction_Price')).alignment = Alignment(
                    horizontal="center"
                )
                ws2.cell(row=13, column=start_column_12).font = Font(bold=True)

            start_column_12 += 2
            end_column_12 += 2

        start_column = start_column_11
        for j in range(4):
            if j == 0 or j == 2:
                ws2.column_dimensions[get_column_letter(start_column)].width = 12
                ws2.cell(row=14, column=start_column, value=label.get(language_code).get('Price_per_PU')).alignment = Alignment(horizontal="left")
                ws2.cell(row=14, column=start_column).font = Font(bold=True)

            if j == 1 or j == 3:
                ws2.column_dimensions[get_column_letter(start_column)].width = 12
                ws2.cell(row=14, column=start_column, value=label.get(language_code).get('Total_price')).alignment = Alignment(horizontal="left")
                ws2.cell(row=14, column=start_column).font = Font(bold=True)

            start_column += 1

        start_column_11 += 4
        end_column_11 += 4
    row_item = 15
    for i, obj in enumerate(items):
        item = obj.get('item')
        ws2.cell(row=row_item, column=1, value=i + 1).alignment = Alignment(horizontal="center")
        ws2.merge_cells(start_row=row_item, start_column=2, end_row=row_item, end_column=3)
        ws2.cell(row=row_item, column=2, value=item.name).alignment = Alignment(horizontal="left")
        ws2.cell(row=row_item, column=4, value=item.unit.name).alignment = Alignment(horizontal="left")

        ws2.column_dimensions[get_column_letter(5)].width = 12
        ws2.cell(row=row_item, column=5, value=item.quantity).alignment = Alignment(horizontal="left")

        ws2.column_dimensions[get_column_letter(6)].width = 12
        ws2.cell(row=row_item, column=6, value=auction.currency.item_code).alignment = Alignment(horizontal="left")

        ws2.column_dimensions[get_column_letter(6)].width = 12
        ws2.cell(row=row_item, column=7, value="{:10,.2f}".format(item.budget)).alignment = Alignment(horizontal="left")

        column_supplier = 10
        for supplier in obj.get("suppliers"):
            ws2.column_dimensions[get_column_letter(column_supplier)].width = 18
            ws2.column_dimensions[get_column_letter(column_supplier + 1)].width = 18
            ws2.cell(row=row_item, column=column_supplier, value="{:10,.2f}".format(supplier.get('price') / item.quantity)).alignment = Alignment(
                horizontal="left"
            )
            ws2.cell(row=row_item, column=column_supplier + 1, value="{:10,.2f}".format(supplier.get('price'))).alignment = Alignment(
                horizontal="left"
            )
            column_supplier += 4

        column_starting_price = 8
        for supplier in obj.get("suppliers"):
            ws2.column_dimensions[get_column_letter(column_starting_price)].width = 18
            ws2.column_dimensions[get_column_letter(column_starting_price + 1)].width = 18
            ws2.cell(
                row=row_item, column=column_starting_price, value="{:10,.2f}".format(supplier.get('confirm_price') / item.quantity)
            ).alignment = Alignment(horizontal="left")
            ws2.cell(row=row_item, column=column_starting_price + 1, value="{:10,.2f}".format(supplier.get('confirm_price'))).alignment = Alignment(
                horizontal="left"
            )
            column_starting_price += 4

        row_item += 1
    max_row = 18
    # case1
    if auction.technical_weighting_id == 1:
        case1(ws2=ws2, row_item=row_item, supplier_sort=supplier_sort, label=label, language_code=language_code)
    # case2
    if auction.technical_weighting_id == 2:
        max_row = 18 + 1
        case2(ws2=ws2, row_item=row_item, supplier_sort=supplier_sort, label=label, language_code=language_code)
    # case3
    if auction.technical_weighting_id == 3:
        max_row = 18 + 1
        case3(ws2=ws2, row_item=row_item, supplier_sort=supplier_sort, label=label, language_code=language_code)

    for col in ws2.iter_cols(min_row=11, max_col=6 + 4 * len(supplier_sort) + 1, max_row=max_row + len(items)):
        for cell in col:
            cell.border = thin_border

    ws_boder = "A11:" + get_column_letter(6 + 4 * len(supplier_sort) + 1) + str(max_row + len(items))
    set_border(ws2, ws_boder)


def auction_results(ws2, auction_result, auction, label, language_code):

    ws2["A5"] = "Auction Result"
    set_background_black(ws=ws2, min_row=5, max_col=3, max_row=5)
    ws2.merge_cells("A5:C5")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col in ws2.iter_cols(min_row=6, max_col=13, max_row=8):
        for cell in col:
            cell.border = thin_border
    ws2.cell(row=6, column=1, value=label.get(language_code).get('Best_Bid')).font = Font(bold=True)
    ws2.cell(row=6, column=4, value=label.get(language_code).get('Saving_Budget_Volume')).font = Font(bold=True)
    ws2.cell(row=6, column=8, value=label.get(language_code).get('Savings_best_1st_offers')).font = Font(bold=True)
    ws2.cell(row=6, column=10, value=label.get(language_code).get('Reach_Target')).font = Font(bold=True)
    ws2.cell(row=6, column=12, value=label.get(language_code).get('Reach_Target_%')).font = Font(bold=True)

    ws2.merge_cells("A6:C6")
    ws2.merge_cells("A7:C8")

    ws2.merge_cells("D6:G6")
    ws2.merge_cells("D7:G7")
    ws2.merge_cells("D8:G8")
    ws2.merge_cells("D7:G7")

    ws2.merge_cells("H6:I6")
    ws2.merge_cells("H7:I7")
    ws2.merge_cells("H8:I8")

    ws2.merge_cells("J6:K6")
    ws2.merge_cells("J7:K8")

    ws2.merge_cells("L6:M6")
    ws2.merge_cells("L7:M7")
    ws2.merge_cells("L8:M8")
    if auction_result.get("best_bid"):
        best_bid = "VND  " + str("{:10,.2f}".format(auction_result.get("best_bid")))
        ws2.cell(row=7, column=1, value=best_bid).alignment = Alignment(vertical="center")
        ws2.cell(row=7, column=1).font = Font(color="32a85e")
        # Savings vs. Budget Volume
        ws2.cell(
            row=7,
            column=4,
            value=str(round(((auction_result["total_budget"] - auction_result.get("best_bid")) / auction_result["total_budget"]) * 100, 2)) + "%",
        ).alignment = Alignment(vertical="center")
        ws2.cell(
            row=8, column=4, value="VND  " + str("{:10,.2f}".format(auction_result["total_budget"] - auction_result.get("best_bid")))
        ).alignment = Alignment(vertical="center")
        # Savings vs. The best 1st offers
        min_confirm_price = min(auction_result.get("list_total_confirm_price_supplier"))
        ws2.cell(
            row=7, column=8, value=str(round(((min_confirm_price - auction_result.get("best_bid")) / min_confirm_price) * 100, 2)) + "%"
        ).alignment = Alignment(vertical="center")
        ws2.cell(row=8, column=8, value="VND  " + str("{:10,.2f}".format(min_confirm_price - auction_result.get("best_bid")))).alignment = Alignment(
            vertical="center"
        )

    supplier_count = AuctionSupplier.objects.filter(auction=auction, is_accept=True, supplier_status__in=[5, 6, 8, 9, 10]).order_by('id').count()
    if auction.auction_type1_id in (1, 3, 4) and supplier_count > 0:
        ws2.cell(
            row=7, column=10, value=str(auction_result.get("count_reach_target")) + label.get(language_code).get('Bidders')
        ).alignment = Alignment(vertical="center")
        ws2.cell(row=7, column=10).font = Font(color="32a85e")
        min_price = auction_result.get("total_target") - auction_result.get("total_reach_target")
        x = round(min_price / auction_result.get("total_target") * 100, 2)
        ws2.cell(row=7, column=12, value=str(x) + "%").alignment = Alignment(vertical="center")
        ws2.cell(row=7, column=12).font = Font(color="32a85e")
        ws2.cell(row=8, column=12, value="{:10,.2f}".format(min_price)).alignment = Alignment(vertical="center")

    set_border(ws2, "A6:M8")


def case1(ws2, row_item, supplier_sort, label, language_code):
    total_bid_list = []
    for supplier in supplier_sort:
        total_bid_list.append(supplier.get("total_bid"))
    total_bid_ranks = numpy.array(total_bid_list).argsort().argsort()

    color_green = "32a85e"
    color_yellow = "FF9933"
    color_red = "FF0000"

    ws2.cell(row=row_item, column=1, value=label.get(language_code).get('Total_Amount_for_Total_Quantity')).alignment = Alignment(horizontal="left")
    ws2.merge_cells(start_row=row_item, start_column=1, end_row=row_item, end_column=7)
    ws2.cell(row=row_item + 1, column=1, value=label.get(language_code).get('Sourcing_Recommendation')).alignment = Alignment(horizontal="left")
    ws2.merge_cells(start_row=row_item + 1, start_column=1, end_row=row_item + 1, end_column=7)
    ws2.cell(row=row_item + 2, column=1, value=label.get(language_code).get('Recommendation_Sourcing_Decision')).alignment = Alignment(
        horizontal="left"
    )
    ws2.merge_cells(start_row=row_item + 2, start_column=1, end_row=row_item + 2, end_column=7)
    ws2.cell(row=row_item + 3, column=1, value=label.get(language_code).get('Others_Comments')).alignment = Alignment(horizontal="left")
    ws2.merge_cells(start_row=row_item + 3, start_column=1, end_row=row_item + 3, end_column=7)

    column = 8
    end_column = 11
    for idx, supplier in enumerate(supplier_sort):
        if total_bid_ranks[idx] == 0:
            rank = label.get(language_code).get('First')
            award = label.get(language_code).get('Award')
            color = color_green
        elif total_bid_ranks[idx] == 1:
            rank = label.get(language_code).get('Second')
            award = None
            color = color_yellow
        elif total_bid_ranks[idx] == 2:
            rank = label.get(language_code).get('Third')
            award = None
            color = color_red
        else:
            rank = None
            award = None
            color = color_red
        column1 = column
        end_column1 = end_column
        for i in range(2):
            if i == 0:
                ws2.cell(row=row_item, column=column1, value="{:10,.2f}".format(supplier.get('total_confirm_price')))
            if i == 1:
                ws2.cell(row=row_item, column=column1, value="{:10,.2f}".format(supplier.get('total_bid'))).font = Font(color=color)
            ws2.cell(row=row_item, column=column1).alignment = Alignment(horizontal="right")
            ws2.merge_cells(start_row=row_item, start_column=column1, end_row=row_item, end_column=column1 + 1)
            column1 += 2

        ws2.cell(row=row_item + 1, column=column, value=rank).font = Font(color=color)
        ws2.cell(row=row_item + 1, column=column).alignment = Alignment(horizontal="right")
        ws2.merge_cells(start_row=row_item + 1, start_column=column, end_row=row_item + 1, end_column=end_column)

        ws2.cell(row=row_item + 2, column=column, value=award).font = Font(color=color)
        ws2.cell(row=row_item + 2, column=column).alignment = Alignment(horizontal="right")
        ws2.merge_cells(start_row=row_item + 2, start_column=column, end_row=row_item + 2, end_column=end_column)

        ws2.cell(row=row_item + 3, column=column, value=None).font = Font(color="32a85e")
        ws2.cell(row=row_item + 3, column=column).alignment = Alignment(horizontal="right")
        ws2.merge_cells(start_row=row_item + 3, start_column=column, end_row=row_item + 3, end_column=end_column)

        column += 4
        end_column += 4


def case2(ws2, row_item, supplier_sort, label, language_code):
    technical_score_list = []
    commercial_against_technical = []
    total_bid_list = []
    for supplier in supplier_sort:
        total_bid_list.append(supplier.get("total_bid"))
        technical_score_list.append(supplier.get("technical_score_supplier"))
        commercial_against_technical.append(round(supplier.get("total_bid") / supplier.get("technical_score_supplier")))
    technical_score_ranks = numpy.array(technical_score_list).argsort().argsort()
    commercial_against_technical_ranks = numpy.array(commercial_against_technical).argsort().argsort()
    total_bid_ranks = numpy.array(total_bid_list).argsort().argsort()

    color_green = "32a85e"
    color_yellow = "FF9933"
    color_red = "FF0000"

    ws2.cell(row=row_item, column=1, value=label.get(language_code).get('Total_Amount_for_Total_Quantity')).alignment = Alignment(horizontal="left")
    ws2.merge_cells(start_row=row_item, start_column=1, end_row=row_item, end_column=7)
    ws2.cell(row=row_item + 1, column=1, value=label.get(language_code).get('Total_technical_point_100')).alignment = Alignment(horizontal="left")
    ws2.merge_cells(start_row=row_item + 1, start_column=1, end_row=row_item + 1, end_column=7)
    ws2.cell(row=row_item + 2, column=1, value=label.get(language_code).get('Commercial_Against_Technical')).alignment = Alignment(horizontal="left")
    ws2.merge_cells(start_row=row_item + 2, start_column=1, end_row=row_item + 2, end_column=7)
    ws2.cell(row=row_item + 3, column=1, value=label.get(language_code).get('Sourcing_Recommendation')).alignment = Alignment(horizontal="left")
    ws2.merge_cells(start_row=row_item + 3, start_column=1, end_row=row_item + 3, end_column=7)
    ws2.cell(row=row_item + 4, column=1, value=label.get(language_code).get('Recommendation_Sourcing_Decision')).alignment = Alignment(
        horizontal="left"
    )
    ws2.merge_cells(start_row=row_item + 4, start_column=1, end_row=row_item + 4, end_column=7)

    column = 8
    end_column = 11
    for idx, supplier in enumerate(supplier_sort):
        if total_bid_ranks[idx] == 0:
            color_total_bid = color_green
        elif total_bid_ranks[idx] == 1:
            color_total_bid = color_yellow
        elif total_bid_ranks[idx] == 2:
            color_total_bid = color_red
        else:
            rank = None
            award = None
            color_total_bid = color_red
        column1 = column
        end_column1 = end_column
        for i in range(2):
            if i == 0:
                ws2.cell(row=row_item, column=column1, value="{:10,.2f}".format(supplier.get('total_confirm_price')))
            if i == 1:
                ws2.cell(row=row_item, column=column1, value="{:10,.2f}".format(supplier.get('total_bid'))).font = Font(color=color_total_bid)
            ws2.cell(row=row_item, column=column1).alignment = Alignment(horizontal="right")
            ws2.merge_cells(start_row=row_item, start_column=column1, end_row=row_item, end_column=column1 + 1)
            column1 += 2
        if technical_score_ranks[idx] == 0:
            color = color_green
        elif technical_score_ranks[idx] == 1:
            color = color_yellow
        else:
            color = color_red
        ws2.cell(row=row_item + 1, column=column, value="{:10,.2f}".format(supplier.get("technical_score_supplier"))).font = Font(color=color)
        ws2.cell(row=row_item + 1, column=column).alignment = Alignment(horizontal="right")
        ws2.merge_cells(start_row=row_item + 1, start_column=column, end_row=row_item + 1, end_column=end_column)
        if commercial_against_technical_ranks[idx] == 0:
            color_rank = color_green
            rank = label.get(language_code).get('First')
            award = label.get(language_code).get('Award')
        elif commercial_against_technical_ranks[idx] == 1:
            color_rank = color_yellow
            rank = label.get(language_code).get('Second')
            award = None
        elif commercial_against_technical_ranks[idx] == 2:
            rank = label.get(language_code).get('Third')
            award = None
            color_rank = color_red
        else:
            rank = None
            award = None
            color_rank = color_red
        ws2.cell(
            row=row_item + 2, column=column, value="{:10,.2f}".format(round(supplier.get("total_bid") / supplier.get("technical_score_supplier")))
        ).font = Font(color=color_rank)
        ws2.cell(row=row_item + 2, column=column).alignment = Alignment(horizontal="right")
        ws2.merge_cells(start_row=row_item + 2, start_column=column, end_row=row_item + 2, end_column=end_column)

        ws2.cell(row=row_item + 3, column=column, value=rank).font = Font(color=color_rank)
        ws2.cell(row=row_item + 3, column=column).alignment = Alignment(horizontal="right")
        ws2.merge_cells(start_row=row_item + 3, start_column=column, end_row=row_item + 3, end_column=end_column)

        ws2.cell(row=row_item + 4, column=column, value=award).font = Font(color=color_rank)
        ws2.cell(row=row_item + 4, column=column).alignment = Alignment(horizontal="right")
        ws2.merge_cells(start_row=row_item + 4, start_column=column, end_row=row_item + 4, end_column=end_column)

        ws2.cell(row=row_item + 5, column=column, value=None).font = Font(color="32a85e")
        ws2.cell(row=row_item + 5, column=column).alignment = Alignment(horizontal="right")
        ws2.merge_cells(start_row=row_item + 5, start_column=column, end_row=row_item + 5, end_column=end_column)
        column += 4
        end_column += 4


def case3(ws2, row_item, supplier_sort, label, language_code):
    total_quantity_after_bonus = []
    total_bid_list = []
    for supplier in supplier_sort:
        total_bid_list.append(supplier.get("total_bid"))
        total_quantity_after_bonus.append(round(supplier.get("total_bid") + supplier.get("technical_score_supplier")))
    total_quantity_after_bonus_ranks = numpy.array(total_quantity_after_bonus).argsort().argsort()
    total_bid_ranks = numpy.array(total_bid_list).argsort().argsort()

    color_green = "32a85e"
    color_yellow = "FF9933"
    color_red = "FF0000"

    ws2.cell(row=row_item, column=1, value=label.get(language_code).get('Total_Amount_for_Total_Quantity')).alignment = Alignment(horizontal="left")
    ws2.merge_cells(start_row=row_item, start_column=1, end_row=row_item, end_column=7)
    ws2.cell(row=row_item + 1, column=1, value=label.get(language_code).get('Bonus_Malus_value')).alignment = Alignment(horizontal="left")
    ws2.merge_cells(start_row=row_item + 1, start_column=1, end_row=row_item + 1, end_column=7)
    ws2.cell(row=row_item + 2, column=1, value=label.get(language_code).get('Total_Amount_for_Total_Quantity_After_Bonus')).alignment = Alignment(
        horizontal="left"
    )
    ws2.merge_cells(start_row=row_item + 2, start_column=1, end_row=row_item + 2, end_column=7)
    ws2.cell(row=row_item + 3, column=1, value=label.get(language_code).get('Sourcing_Recommendation')).alignment = Alignment(horizontal="left")
    ws2.merge_cells(start_row=row_item + 3, start_column=1, end_row=row_item + 3, end_column=7)
    ws2.cell(row=row_item + 4, column=1, value=label.get(language_code).get('Recommendation_Sourcing_Decision')).alignment = Alignment(
        horizontal="left"
    )
    ws2.merge_cells(start_row=row_item + 4, start_column=1, end_row=row_item + 4, end_column=7)

    column = 8
    end_column = 11
    for idx, supplier in enumerate(supplier_sort):
        if total_bid_ranks[idx] == 0:
            color_total_bid = color_green
        elif total_bid_ranks[idx] == 1:
            color_total_bid = color_yellow
        elif total_bid_ranks[idx] == 2:
            color_total_bid = color_red
        else:
            color_total_bid = color_red
        column1 = column
        end_column1 = end_column
        for i in range(2):
            if i == 0:
                ws2.cell(row=row_item, column=column1, value="{:10,.2f}".format(supplier.get('total_confirm_price')))
            if i == 1:
                ws2.cell(row=row_item, column=column1, value="{:10,.2f}".format(supplier.get('total_bid'))).font = Font(color=color_total_bid)
            ws2.cell(row=row_item, column=column1).alignment = Alignment(horizontal="right")
            ws2.merge_cells(start_row=row_item, start_column=column1, end_row=row_item, end_column=column1 + 1)
            column1 += 2
        if supplier.get("technical_score_supplier") > 0:
            color = color_red
        else:
            color = color_green
        ws2.cell(row=row_item + 1, column=column, value=supplier.get("technical_score_supplier")).font = Font(color=color)
        ws2.cell(row=row_item + 1, column=column).alignment = Alignment(horizontal="right")
        ws2.merge_cells(start_row=row_item + 1, start_column=column, end_row=row_item + 1, end_column=end_column)

        if total_quantity_after_bonus_ranks[idx] == 0:
            rank = label.get(language_code).get('First')
            award = label.get(language_code).get('Award')
            color_rank = color_green
        elif total_quantity_after_bonus_ranks[idx] == 1:
            rank = label.get(language_code).get('Second')
            award = None
            color_rank = color_yellow
        elif total_quantity_after_bonus_ranks[idx] == 2:
            color_rank = color_red
            rank = label.get(language_code).get('Third')
            award = None
        else:
            rank = None
            award = None
            color_rank = color_red
        ws2.cell(
            row=row_item + 2, column=column, value="{:10,.2f}".format(round(supplier.get("total_bid") + supplier.get("technical_score_supplier")))
        ).font = Font(color=color_rank)
        ws2.cell(row=row_item + 2, column=column).alignment = Alignment(horizontal="right")
        ws2.merge_cells(start_row=row_item + 2, start_column=column, end_row=row_item + 2, end_column=end_column)

        ws2.cell(row=row_item + 3, column=column, value=rank).font = Font(color=color_rank)
        ws2.cell(row=row_item + 3, column=column).alignment = Alignment(horizontal="right")
        ws2.merge_cells(start_row=row_item + 3, start_column=column, end_row=row_item + 3, end_column=end_column)

        ws2.cell(row=row_item + 4, column=column, value=award).font = Font(color=color_rank)
        ws2.cell(row=row_item + 4, column=column).alignment = Alignment(horizontal="right")
        ws2.merge_cells(start_row=row_item + 4, start_column=column, end_row=row_item + 4, end_column=end_column)

        ws2.cell(row=row_item + 5, column=column, value=None).font = Font(color="32a85e")
        ws2.cell(row=row_item + 5, column=column).alignment = Alignment(horizontal="right")
        ws2.merge_cells(start_row=row_item + 5, start_column=column, end_row=row_item + 5, end_column=end_column)
        column += 4
        end_column += 4


def sheet3(wb, auction, language_code):
    tz = pytz.timezone(auction.user.local_time)
    start_time = auction.start_time.replace(tzinfo=pytz.utc).astimezone(tz)
    start_time = datetime.strftime(start_time, '%d-%b-%Y %H:%M:%S')

    contract_type = auction.contract_type
    if ContractTypeTranslation.objects.filter(contract_type=contract_type, language_code=language_code).exists():
        contract_type = ContractTypeTranslation.objects.filter(contract_type=contract_type, language_code=language_code).first()
    name = contract_type.name
    if auction.contract_type_id == 2:
        contract_type = (
            auction.contract_type.name
            + " "
            + datetime.strftime(auction.contract_from, '%d.%m.%Y')
            + " - "
            + datetime.strftime(auction.contract_to, '%d.%m.%Y')
        )
    else:
        contract_type = auction.contract_type.name
    delivery_address = auction.delivery_address
    delivery_term = auction.delivery_term
    if DeliveryTermTranslation.objects.filter(delivery_term=delivery_term, language_code=language_code).exists():
        delivery_term = DeliveryTermTranslation.objects.filter(delivery_term=delivery_term, language_code=language_code).first()
    delivery_term = delivery_term.name
    note_from_buyer = auction.note_from_buyer

    auction_item = AuctionItem.objects.filter(auction=auction)
    auction_supplier = AuctionSupplier.objects.filter(auction=auction, is_accept=True, supplier_status__in=[5, 6, 8, 9, 10]).order_by('id')
    supplier_list = []

    for supplier in auction_supplier:
        is_accept_rule = supplier.is_accept_rule
        supplier_bids = AuctionBid.objects.filter(user=supplier.user, auction=auction).order_by("id")
        result = {"auction_supplier": supplier, "supplier_bids": supplier_bids}
        items = []
        total_bid = 0
        for item in auction_item:
            auction_bids = AuctionBid.objects.filter(user=supplier.user, auction_item=item).order_by('id')
            auction_item_supplier = AuctionItemSupplier.objects.filter(auction_item=item, auction_supplier=supplier).first()
            if auction_bids.exists():
                price = auction_bids.last().price
            else:
                price = auction_item_supplier.confirm_price
            total_bid += price
            items.append(
                {'item': item, 'bid_item': price, 'confirm_price': auction_item_supplier.confirm_price, 'price': auction_item_supplier.price}
            )
            result["items"] = items
        result["total_bid"] = total_bid
        supplier_list.append(result)

    label = {
        "vi": {
            "Description_type": "STT/Mô tả",
            "Offer": "Bên mời thầu",
            "Offer_number": "Mã nhà thầu:",
            "Date": "Ngày:",
            "Supplier": "Nhà cung cấp:",
            "Auction_Rule": "Quy tắc đấu giá",
            "General_terms_and_conditions": "Các điều khoản và điều kiện chung:",
            "Award_of_contract_process_after_Auction": "Trao thầu của hợp đồng sau khi đấu giá:",
            "Slip_Order_Possibility": "Khả năng trượt thầu:",
            "Contract_Type": "Loại hợp đồng:",
            "Delivery_term": "Điều khoản giao hàng:",
            "Delivery_address": "Địa chỉ giao hàng:",
            "Notes_from_Buyer": "Ghi chú từ người mua:",
            "Acceptance_of_auction_rule": "Sự chấp thuận của quy tắc đấu giá",
            "Item": "Mục",
            "overview": "Tổng quát",
            "Currency": "Đơn vị tiền tệ:",
            "Item_name": "Tên mặt hàng",
            "Specification": "Thông số kỹ thuật",
            "Quantity": "Số lượng",
            "Info_fields_buyers_only": "Trường thông tin (chỉ người mua)",
            "Starting_price": "Giá khởi điểm",
            "Price": "Giá",
            "Total": "Tổng cộng",
            "Total_price_of_all_items": "Tổng giá của tất cả các mặt hàng",
            "Report_Bid": "Báo cáo dự thầu",
            "Time": "Thời gian",
            "Total_number_bid": " Tổng số dự thầu:",
            "Price_per_PU": "Giá mỗi [PU]",
            "Price_per_position": "Giá mỗi vị trí",
            "Comment": "Ghi chú",
        },
        "en": {
            "Description_type": "No./Description of type",
            "Offer": "Offer",
            "Offer_number": "Offer number:",
            "Date": "Date:",
            "Supplier": "Supplier:",
            "Auction_Rule": "Auction Rule",
            "General_terms_and_conditions": "General terms and conditions:",
            "Award_of_contract_process_after_Auction": "Award of contract process after Auction:",
            "Slip_Order_Possibility": "Slip Order Possibility:",
            "Contract_Type": "Contract Type:",
            "Delivery_term": "Delivery term: ",
            "Delivery_address": "Delivery address:",
            "Notes_from_Buyer": "Notes from Buyer:",
            "Acceptance_of_auction_rule": "Acceptance of auction rule::",
            "Item": "Item ",
            "overview": "overview",
            "Currency": "Currency",
            "Item_name": "Item name",
            "Specification": "Specification",
            "Quantity": "Quantity",
            "Info_fields_buyers_only": "Info fields (buyers only)",
            "Starting_price": "Starting price",
            "Price": "Price",
            "Total": " Total",
            "Total_price_of_all_items": " Total price of all items",
            "Report_Bid": " Report Bid",
            "Total_number_bid": " Total number bid:",
            "Time": " Time",
            "Price_per_PU": "Price per [PU]",
            "Price_per_position": "Price per position",
            "Comment": "Comment",
        },
    }

    for i, supplier in enumerate(supplier_list):
        sheet_number = 3 + i
        sheet = "wb" + str(sheet_number)
        sheet = wb.create_sheet("Supplier " + f'''{supplier.get('auction_supplier').user.full_name}''', 2 + i)
        sheet.sheet_view.showGridLines = False

        sheet.merge_cells("A2:C3")
        sheet.cell(row=2, column=1, value=label.get(language_code).get('Description_type')).alignment = Alignment(vertical="center")
        sheet['D2'] = f'''{auction.item_code} - {auction.title} '''
        sheet['D3'] = str(auction.auction_type1)

        img = Image(os.path.join(settings.STATICFILES_DIRS[0], 'logo_report.png'))
        img.height = 40
        img.width = 300
        sheet.column_dimensions["D"].width = 20.0

        sheet.merge_cells("F2:I3")
        sheet.add_image(img, 'H2')
        sheet.column_dimensions["C"].width = 18.0
        set_background_black(ws=sheet, min_row=2, max_col=12, max_row=3)

        # Offer
        sheet["A5"] = label.get(language_code).get('Offer')
        sheet.cell(row=5, column=1, value=label.get(language_code).get('Offer')).font = Font(bold=True)
        sheet.cell(row=7, column=1, value=label.get(language_code).get('Offer_number')).font = Font(bold=True)
        sheet.cell(row=9, column=1, value=label.get(language_code).get('Date')).font = Font(bold=True)
        sheet.cell(row=11, column=1, value=label.get(language_code).get('Supplier')).font = Font(bold=True)
        sheet.cell(row=7, column=4, value=str(auction.item_code) + "-" + str(i + 1)).font = Font()
        sheet.cell(row=11, column=4, value=supplier.get("auction_supplier").user.supplier.company_full_name).font = Font()
        sheet.cell(row=9, column=4, value=str(start_time)).font = Font()

        set_background_black(ws=sheet, min_row=5, max_col=3, max_row=5)
        sheet.merge_cells("A5:C5")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        set_border(ws=sheet, cell_range=("A6:L12"))

        # Auction Rule
        sheet["A14"] = label.get(language_code).get('Auction_Rule')
        sheet.cell(row=16, column=1, value=label.get(language_code).get('General_terms_and_conditions')).font = Font(bold=True)
        sheet.cell(row=18, column=1, value=label.get(language_code).get('Award_of_contract_process_after_Auction')).font = Font(bold=True)
        sheet.cell(row=21, column=1, value=label.get(language_code).get('Slip_Order_Possibility')).font = Font(bold=True)
        sheet.cell(row=23, column=1, value=label.get(language_code).get('Contract_Type')).font = Font(bold=True)
        sheet.cell(row=25, column=1, value=label.get(language_code).get('Delivery_term')).font = Font(bold=True)
        sheet.cell(row=27, column=1, value=label.get(language_code).get('Delivery_address')).font = Font(bold=True)
        sheet.cell(row=29, column=1, value=label.get(language_code).get('Notes_from_Buyer')).font = Font(bold=True)
        sheet.cell(row=31, column=1, value=label.get(language_code).get('Acceptance_of_auction_rule')).font = Font(bold=True)
        sheet["D16"].hyperlink = "https://nextpro.io/term-conditions/"
        sheet.cell(row=16, column=4, value="https://nextpro.io/term-conditions/").font = Font()
        sheet.cell(row=23, column=4, value=contract_type).font = Font()
        sheet.cell(row=25, column=4, value=delivery_term).font = Font()
        sheet.cell(row=27, column=4, value=delivery_address).font = Font()
        sheet.cell(row=29, column=4, value=note_from_buyer).font = Font()
        sheet.cell(row=31, column=4, value="").font = Font()

        set_background_black(ws=sheet, min_row=14, max_col=3, max_row=14)
        sheet.merge_cells("A14:C14")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        set_border(ws=sheet, cell_range=("A15:L31"))

        # Item overview
        sheet["A34"] = label.get(language_code).get('Item')
        sheet['B34'] = label.get(language_code).get('overview')
        sheet.cell(row=35, column=1, value="Pos.").font = Font(bold=True)
        sheet.cell(row=35, column=2, value=label.get(language_code).get('Item_name')).font = Font(bold=True)
        sheet.cell(row=35, column=7, value=label.get(language_code).get('Quantity')).font = Font(bold=True)
        sheet.column_dimensions[get_column_letter(10)].width = 18
        sheet.cell(row=35, column=10, value=label.get(language_code).get('Price_per_PU')).font = Font(bold=True)
        sheet.column_dimensions[get_column_letter(11)].width = 18
        sheet.cell(row=35, column=11, value=label.get(language_code).get('Price_per_position')).font = Font(bold=True)
        sheet.column_dimensions[get_column_letter(12)].width = 12
        sheet.cell(row=35, column=12, value=label.get(language_code).get('Comment')).font = Font(bold=True)
        row = 36
        column = 1
        thin_border = Border(left=None, right=None, top=None, bottom=Side(style='thin'))
        for i, item in enumerate(supplier.get("items")):
            sheet.cell(row=row, column=column, value=i + 1).font = Font(bold=True)
            sheet.cell(row=row, column=column + 1, value=item.get("item").name).font = Font()
            sheet.cell(row=row, column=column + 3, value=label.get(language_code).get('Specification')).font = Font(bold=True)
            sheet.column_dimensions[get_column_letter(column + 3)].width = 25
            sheet.cell(row=row + 1, column=column + 3, value=label.get(language_code).get('Info_fields_buyers_only')).font = Font(bold=True)
            sheet.column_dimensions[get_column_letter(column + 4)].width = 14
            sheet.cell(row=row + 1, column=column + 4, value=label.get(language_code).get('Starting_price')).font = Font(bold=True)
            sheet.cell(row=row + 2, column=column + 4, value="{:10,.2f}".format(item.get("confirm_price"))).font = Font()

            sheet.column_dimensions[get_column_letter(column + 5)].width = 14
            sheet.cell(row=row + 1, column=column + 5, value=label.get(language_code).get('Price')).font = Font(bold=True)
            sheet.cell(row=row + 2, column=column + 5, value="{:10,.2f}".format(item.get("price"))).font = Font()

            sheet.cell(row=row + 1, column=column + 6, value=item.get("item").quantity).font = Font()
            sheet.cell(row=row + 2, column=column + 8, value=label.get(language_code).get('Total')).font = Font(bold=True)
            sheet.cell(row=row + 2, column=column + 7, value=1).font = Font(bold=True)
            sheet.cell(row=row + 2, column=column + 9, value="{:10,.2f}".format(item.get("bid_item") / item.get("item").quantity))
            sheet.cell(row=row + 2, column=column + 10, value="{:10,.2f}".format(item.get("bid_item"))).font = Font()
            for x in range(12):
                sheet.cell(row=row + 2, column=x + 1).border = thin_border
            row += 3

        sheet.cell(row=row, column=4, value=label.get(language_code).get('Total_price_of_all_items')).font = Font(bold=True)
        sheet.cell(row=row, column=11, value="{:10,.2f}".format(supplier.get("total_bid"))).font = Font()
        set_background_black(ws=sheet, min_row=34, max_col=3, max_row=34)
        sheet.merge_cells(start_row=34, start_column=9, end_row=34, end_column=12)
        sheet.cell(row=34, column=9).font = Font(size=13, bold=True, color="ffffff")
        currency = auction.currency
        if CurrencyTranslation.objects.filter(currency=currency, language_code=language_code).exists():
            currency = CurrencyTranslation.objects.filter(currency=currency, language_code=language_code).first()
        currency = currency.name

        sheet.cell(row=34, column=9, value=label.get(language_code).get('Currency') + currency).fill = PatternFill(
            fgColor="000000", fill_type='solid'
        )
        set_border(ws=sheet, cell_range=("A35:L" + str(row)))

        # report bid
        sheet.cell(row=row + 2, column=1, value=label.get(language_code).get('Report_Bid')).font = Font(bold=True)
        set_background_black(ws=sheet, min_row=row + 2, max_col=3, max_row=row + 2)
        sheet.cell(row=row + 3, column=1, value="No.").font = Font(bold=True)
        sheet.merge_cells(start_row=row + 3, start_column=2, end_row=row + 3, end_column=3)
        sheet.cell(row=row + 3, column=2, value=label.get(language_code).get('Item_name')).font = Font(bold=True)
        sheet.merge_cells(start_row=row + 3, start_column=4, end_row=row + 3, end_column=5)
        sheet.cell(row=row + 3, column=4, value=label.get(language_code).get('Time')).font = Font(bold=True)
        sheet.merge_cells(start_row=row + 3, start_column=6, end_row=row + 3, end_column=7)
        sheet.cell(row=row + 3, column=6, value=label.get(language_code).get('Price')).font = Font(bold=True)
        row_bid = row + 4
        for j, bid in enumerate(supplier.get("supplier_bids")):
            sheet.cell(row=row_bid, column=1, value=j + 1).font = Font()
            sheet.merge_cells(start_row=row_bid, start_column=2, end_row=row_bid, end_column=3)
            sheet.cell(row=row_bid, column=2, value=bid.auction_item.name).font = Font()
            sheet.merge_cells(start_row=row_bid, start_column=4, end_row=row_bid, end_column=5)
            created = bid.created.replace(tzinfo=pytz.utc).astimezone(tz)
            created = datetime.strftime(created, '%d-%b-%Y %H:%M:%S')
            sheet.cell(row=row_bid, column=4, value=created).font = Font()
            sheet.merge_cells(start_row=row_bid, start_column=6, end_row=row_bid, end_column=7)
            price = "{:10,.2f}".format(bid.price)
            sheet.cell(row=row_bid, column=6, value=price).font = Font()
            row_bid += 1
        # total number bid
        sheet.cell(row=row_bid, column=1, value=label.get(language_code).get('Total_number_bid')).font = Font(bold=True)
        total_number_bid = label.get(language_code).get('Total_number_bid')
        if supplier.get("supplier_bids"):
            total_number_bid = total_number_bid + str(len(supplier.get("supplier_bids")))
        sheet.cell(row=row_bid, column=1, value=total_number_bid).font = Font(bold=True)
        sheet.merge_cells(start_row=row_bid, start_column=1, end_row=row_bid, end_column=7)
        # table
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for col in sheet.iter_cols(min_row=row + 3, max_col=7, max_row=row_bid):
            for cell in col:
                cell.border = thin_border
        set_border(ws=sheet, cell_range=("A" + str(row + 3) + ":G" + str(row_bid)))


# set layout
def set_border(ws, cell_range):
    rows = ws[cell_range]
    side = Side(border_style='medium', color="FF000000")

    rows = list(rows)
    max_y = len(rows) - 1
    for pos_y, cells in enumerate(rows):
        max_x = len(cells) - 1
        for pos_x, cell in enumerate(cells):
            border = Border(left=cell.border.left, right=cell.border.right, top=cell.border.top, bottom=cell.border.bottom)
            if pos_x == 0:
                border.left = side
            if pos_x == max_x:
                border.right = side
            if pos_y == 0:
                border.top = side
            if pos_y == max_y:
                border.bottom = side

            if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                cell.border = border


def set_background_black(ws, min_row, max_col, max_row):
    for col in ws.iter_cols(min_row=min_row, max_col=max_col, max_row=max_row):
        for cell in col:
            cell.font = Font(size=13, bold=True, color="ffffff")
            cell.fill = PatternFill(fgColor="000000", fill_type='solid')
