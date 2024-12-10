import csv
import json 
import os
import uuid
import openpyxl
from datetime import date

from apps.master_data.models import Promotion, PromotionUserUsed, Position
from apps.users.models import Supplier, Buyer, SupplierCategory, UsersPermission, BuyerActivity, SupplierActivity
from apps.users.serializers import (
    SupplierListSerializer,
    BuyerSerializer,
    SupplierCategorySerializer,
    AdminPermissionSerializer,
    BuyerExportSerializer,
    BuyerActivitySerializer,
    SupplierExportSerializer,
    SupplierActivitySerializer,
    PromotionExportSerializer,
    PromotionUserUsedExportSerializer,
)
from django.conf import settings
from django.contrib.auth import authenticate, get_user_model
from django.http import HttpResponse, Http404

from rest_framework import generics, views, serializers, parsers
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter
from docxtpl import DocxTemplate
User = get_user_model()

class UserProfileSerializer(serializers.ModelSerializer):
    class Meta:
        model = User
        fields = ('id', 'username', 'user_type', 'email', 'first_name', 'last_name')
        read_only_fields = (
            'id',
            'username',
            'user_type',
            'email',
        )


class BuyerProfileSerializer(serializers.ModelSerializer):
    company_logo = serializers.ImageField(max_length=None, use_url=True, allow_null=True, required=False)
    picture = serializers.ImageField(max_length=None, use_url=True, allow_null=True, required=False)

    class Meta:
        model = Buyer
        fields = (
            'id',
            'company_email',
            'company_short_name',
            'company_long_name',
            'company_logo',
            'company_tax',
            'company_address',
            'company_city',
            'company_country',
            'company_country_state',
            'company_number_of_employee',
            'company_website',
            'company_referral_code',
            'gender',
            'picture',
            'phone',
            'position',
            'level',
        )
        read_only_fields = ('id',)


class SupplierProfileSerializer(serializers.ModelSerializer):
    company_logo = serializers.ImageField(max_length=None, use_url=True, allow_null=True, required=False)

    class Meta:
        model = Supplier
        fields = (
            'id',
            'company_short_name',
            'company_long_name',
            'company_tax',
            'company_logo',
            'company_address',
            'company_city',
            'company_country',
            'company_country_state',
            'company_ceo_owner_name',
            'company_ceo_owner_email',
            'company_number_of_employee',
            'company_website',
            'company_credential_profile',
            'gender',
            'phone',
            'position',
            'level',
            'bank_name',
            'bank_code',
            'bank_address',
            'beneficiary_name',
            'switch_bic_code',
            'bank_account_number',
            'bank_currency',
            'international_bank',
            'supplier_form_registration',
            'bank_certification',
            'quality_certification',
            'business_license',
            'tax_certification',
        )
        read_only_fields = ('id',)


class ProfileView(views.APIView):

    permission_classes = [IsAuthenticated]
    parsers_classes = [parsers.MultiPartParser, parsers.FormParser]

    def get(self, request, format=None):

        usersSerializer = UserProfileSerializer(request.user)

        if request.user.isBuyer():
            profileSerializer = BuyerProfileSerializer(request.user.get_profile(), context={'request': request})

            return Response({**usersSerializer.data, **profileSerializer.data,})

        elif request.user.isSupplier():
            profileSerializer = SupplierProfileSerializer(request.user.get_profile(), context={'request': request})

            supplier_categories = SupplierCategory.objects.filter(user_supplier_id=profileSerializer.data['id'])
            SupplierCategorySerializer = SupplierCategorySerializer(supplier_categories, many=True)

            return Response({**usersSerializer.data, **profileSerializer.data, 'categories': SupplierCategorySerializer.data,})
        elif request.user.isAdmin():
            return Response("")

    def put(self, request, format=None):

        usersSerializer = UserProfileSerializer(request.user, data=request.data)

        if usersSerializer.is_valid():

            user = usersSerializer.save()

            if request.user.isBuyer():
                profileSerializer = BuyerProfileSerializer(request.user.get_profile(), data=request.data)
                if profileSerializer.is_valid():
                    profileSerializer.save()

                    return Response({'success': True})
                else:
                    return Response(profileSerializer.errors)

            elif request.user.isSupplier():
                profileSerializer = SupplierProfileSerializer(request.user.get_profile(), data=request.data)

                if profileSerializer.is_valid():
                    profileSerializer.save()

                    supplier_categories = SupplierCategory.objects.filter(user_supplier_id=profileSerializer.data['id'])

                    categories_mapping = {category.category_id: category for category in supplier_categories}

                    categories = json.loads(request.data.get('categories', '[]'))

                    data_mapping = {category['category_id']: category for category in categories}

                    total_percentage = 0
                    for supplier_category in categories:
                        total_percentage = total_percentage + supplier_category['percentage']

                    if total_percentage > 100:
                        return Response({'error': 'Total percentage of all categories chosen cannot be greater than 100%'})

                    else:
                        for category_id, data in data_mapping.items():
                            category = categories_mapping.get(category_id, None)

                            SupplierCategory = SupplierCategorySerializer(category, data=data)
                            if SupplierCategory.is_valid():
                                SupplierCategory.save()
                            else:
                                return Response(SupplierCategory.errors)

                        for category_id, category in categories_mapping.items():
                            if category_id not in data_mapping:
                                category.delete()

                    return Response(profileSerializer.data)
                else:
                    return Response(profileSerializer.errors)

            return Response({'success': True})
        else:
            return Response(usersSerializer.errors)


class ChangePasswordView(views.APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request, format=None):

        password = request.data.get('password', '')
        new_password = request.data.get('new_password', '')
        confirm_password = request.data.get('confirm_password', '')

        if new_password != confirm_password:
            return Response({'error': 'Password and Confirm password are not match'})

        if len(new_password) < 6:
            return Response({'error': 'Password must be at least 6 characters'})

        user = authenticate(username=request.user.username, password=password)

        if not user:
            return Response({'error': 'Old Password is incorrect'})

        user.set_password(new_password)
        user.save()

        return Response({'success': True})


class SupplierList(generics.ListAPIView):
    pagination_class = None
    queryset = User.objects.select_related('supplier').all()
    serializer_class = SupplierListSerializer

    def get(self, request, format=None):
        users = User.objects.filter(user_type=3).select_related('supplier')
        serializers = SupplierListSerializer(users, many=True, context={'request': self.request})
        return Response(serializers.data)

class BuyerList(generics.ListAPIView):
    pagination_class = None
    queryset = Buyer.objects.all()
    serializer_class = BuyerSerializer

class AdminPermissionExport(APIView):
    def get(self, request, format=None):
        permissions = UsersPermission.objects.all().order_by('id').exclude(permission__role=1)
        serializer = AdminPermissionSerializer(permissions, many=True)
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="admin_permission.csv"'
        response.write(u'\ufeff'.encode('utf8'))
        writer = csv.DictWriter(
            response,
            fieldnames=[
                'Change Code',
                'Account ID',
                'Account Name',
                'Account email',
                'Date created',
                'Valid From',
                'Valid to',
                'Roles',
                'Modules',
                'Status',
            ],
        )
        writer.writeheader()
        for permission in serializer.data:
            writer.writerow(
                {
                    'Change Code': permission.get('id'),
                    'Account ID': permission.get('username'),
                    'Account Name': permission.get('short_name'),
                    'Account email': permission.get('email'),
                    'Date created': permission.get('created'),
                    'Valid From': permission.get('valid_from'),
                    'Valid to': permission.get('valid_to'),
                    'Roles': permission.get('role'),
                    'Modules': permission.get('modules'),
                    'Status': permission.get('status'),
                }
            )

        return response


class BuyerExport(APIView):
    def get(self, request, format=None):
        buyer = Buyer.objects.all().order_by('id')
        buyer_list = BuyerExportSerializer(buyer, many=True)
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="buyer_list.csv"'
        response.write(u'\ufeff'.encode('utf8'))
        writer = csv.DictWriter(
            response,
            fieldnames=[
                'Log ID',
                'Account ID',
                'Account email',
                'Date created',
                'Valid From',
                'Valid to',
                'Profile Features',
                'Auctions/year',
                'RFx Cancellation',
                'Status',
                'Changed By',
                'Change Date',
                'Reason In Manual',
            ],
        )
        writer.writeheader()

        for buyer in buyer_list.data:
            buyer_activity_list = BuyerActivity.objects.filter(buyer_id=buyer.get('id')).order_by('id')
            buyer_activities = BuyerActivitySerializer(buyer_activity_list, many=True)
            if len(buyer_activities.data) > 0:
                for buyer_activity in buyer_activities.data:
                    writer.writerow(
                        {
                            'Log ID': buyer.get('id'),
                            'Account ID': buyer.get('username'),
                            'Account email': buyer.get('email'),
                            'Date created': buyer.get('created'),
                            'Valid From': buyer.get('valid_from'),
                            'Valid to': buyer.get('valid_to'),
                            'Profile Features': buyer.get('profile_features'),
                            'Auctions/year': buyer.get('no_eauction_year'),
                            'RFx Cancellation': buyer.get('rfx_year'),
                            'Status': buyer.get('status'),
                            'Changed By': buyer_activity.get('changed_by'),
                            'Change Date': buyer_activity.get('changed_date'),
                            'Reason In Manual': buyer_activity.get('reason_manual'),
                        }
                    )
            else:
                writer.writerow(
                    {
                        'Log ID': buyer.get('id'),
                        'Account ID': buyer.get('username'),
                        'Account email': buyer.get('email'),
                        'Date created': buyer.get('created'),
                        'Valid From': buyer.get('valid_from'),
                        'Valid to': buyer.get('valid_to'),
                        'Profile Features': buyer.get('profile_features'),
                        'Auctions/year': buyer.get('no_eauction_year'),
                        'RFx Cancellation': buyer.get('rfx_year'),
                        'Status': buyer.get('status'),
                    }
                )
        return response


class SupplierExport(APIView):
    def get(self, request, format=None):
        supplier = Supplier.objects.all().order_by('id')
        supplier_list = SupplierExportSerializer(supplier, many=True)
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="supplier_list.csv"'
        response.write(u'\ufeff'.encode('utf8'))
        writer = csv.DictWriter(
            response,
            fieldnames=[
                'Log ID',
                'Account ID',
                'Account email',
                'Date created',
                'Valid From',
                'Valid to',
                'Profile Features',
                'Report/year',
                'Flashsale',
                'Status',
                'Changed By',
                'Change Date',
                'Reason In Manual',
            ],
        )
        writer.writeheader()

        for supplier in supplier_list.data:
            supplier_activity_list = SupplierActivity.objects.filter(supplier_id=supplier.get('id')).order_by('id')
            supplier_activities = SupplierActivitySerializer(supplier_activity_list, many=True)
            if len(supplier_activities.data) > 0:
                for supplier_activity in supplier_activities.data:
                    writer.writerow(
                        {
                            'Log ID': supplier.get('id'),
                            'Account ID': supplier.get('username'),
                            'Account email': supplier.get('email'),
                            'Date created': supplier.get('created'),
                            'Valid From': supplier.get('valid_from'),
                            'Valid to': supplier.get('valid_to'),
                            'Profile Features': supplier.get('profile_features'),
                            'Report/year': supplier.get('report_year'),
                            'Flashsale': supplier.get('flash_sale'),
                            'Status': supplier.get('status'),
                            'Changed By': supplier_activity.get('changed_by'),
                            'Change Date': supplier_activity.get('changed_date'),
                            'Reason In Manual': supplier_activity.get('reason_manual'),
                        }
                    )
            else:
                writer.writerow(
                    {
                        'Log ID': supplier.get('id'),
                        'Account ID': supplier.get('username'),
                        'Account email': supplier.get('email'),
                        'Date created': supplier.get('created'),
                        'Valid From': supplier.get('valid_from'),
                        'Valid to': supplier.get('valid_to'),
                        'Profile Features': supplier.get('profile_features'),
                        'Report/year': supplier.get('report_year'),
                        'Flashsale': supplier.get('flash_sale'),
                        'Status': supplier.get('status'),
                    }
                )
        return response


class DownloadSupplierForm(APIView):
    def get(self, request):
        file_path = os.path.join(settings.STATICFILES_DIRS[0], "supplier_registration.xlsx")
        if os.path.exists(file_path):
            with open(file_path, 'rb') as fh:
                response = HttpResponse(fh.read(), content_type=" application/vnd.openxmlformats-officedocument.wordprocessingml.document")
                response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
                return response
        raise Http404

class PromotionExport(APIView):
    def get(self, request, format=None):
        promotion = Promotion.objects.all().order_by('id')
        promotion_list = PromotionExportSerializer(promotion, many=True)
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="promotion_list.csv"'
        response.write(u'\ufeff'.encode('utf8'))
        writer = csv.DictWriter(
            response,
            fieldnames=[
                'ID',
                'Promotion Program',
                'Description(EN)',
                'Description(VI)',
                'User Receiver', 
                'Email Receiver', 
                'Commission (%)',
                'Discount (%)',
                'Valid From',
                'Valid To',               
                'Applied For Buyer',
                'Applied For Supplier',
                'Status',
                'Visible',                
            ],
        )
        writer.writeheader()

        for promotion in promotion_list.data:
            writer.writerow(
                {
                    'ID': promotion.get('id'),
                    'Promotion Program': promotion.get('name'),
                    'Description(EN)': promotion.get('description'),
                    'Description(VI)': promotion.get('translated'),
                    'User Receiver': promotion.get('user_given'),
                    'Email Receiver': promotion.get('user_given_email'),
                    'Commission (%)': promotion.get('commission'),
                    'Discount (%)': promotion.get('discount'),
                    'Valid From': promotion.get('valid_from'),
                    'Valid To': promotion.get('valid_to'),                   
                    'Applied For Buyer': promotion.get('apply_for_buyer'),
                    'Applied For Supplier': promotion.get('apply_for_supplier'),
                    'Status': promotion.get('status'),
                    'Visible': promotion.get('visible'),                    
                }
            )
        return response

class PromotionUserUsedExport(APIView):
    def get(self, request, format=None):
        promotion_user = PromotionUserUsed.objects.all().order_by('id')
        promotion_user_list = PromotionUserUsedExportSerializer(promotion_user, many=True)
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="promotion_history_list.csv"'
        response.write(u'\ufeff'.encode('utf8'))
        writer = csv.DictWriter(
            response,
            fieldnames=[
                'ID',
                'Title',
                'User Used Name',
                'User Used Code',
                'User Used Email', 
                'Date Used', 
                'Real Amount',
                'Amount After Discount',
                'Promotion Program',
                'Valid From',
                'Valid To',               
                'Visible',
                'User Given',
                'User Given Email',
                'Commission (%)',
                'Status',
                'Discount (%)',   
            ],
        )
        writer.writeheader()

        for promotion_user in promotion_user_list.data:
            writer.writerow(
                {
                    'ID': promotion_user.get('id'),
                    'Title': promotion_user.get('title'),
                    'User Used Name': promotion_user.get('user_name'),
                    'User Used Code': promotion_user.get('user_used'),
                    'User Used Email': promotion_user.get('user_used_email'),
                    'Date Used': promotion_user.get('date_used'),
                    'Real Amount': promotion_user.get('real_amount'),
                    'Amount After Discount': promotion_user.get('amount_after_discount'),
                    'Promotion Program': promotion_user.get('name'),
                    'Valid From': promotion_user.get('valid_from'),
                    'Valid To': promotion_user.get('valid_to'),
                    'Visible': promotion_user.get('visible'),                   
                    'User Given': promotion_user.get('user_given'),
                    'User Given Email': promotion_user.get('user_given_email'),
                    'Commission (%)': promotion_user.get('commission'),
                    'Status': promotion_user.get('status'),
                    'Discount (%)': promotion_user.get('discount'),                                        
                }
            )
        return response


class DownloadSupplierCooperationAgreement(APIView):
    def post(self, request, format=None):
        company_full_name = request.data.get('companyFullName')
        company_address = request.data.get('companyAddress')
        company_tax_code = request.data.get('companyTax')
        company_phone_number = request.data.get('phone')
        contact_person_name = request.data.get('fullName')
        contact_person_position = request.data.get('position')
        lang = request.data.get('lang')

        if company_full_name is None or len(company_full_name) == 0:
            return Http404
        if company_address is None or len(company_address) == 0:
            return Http404
        if company_tax_code is None or len(company_tax_code) == 0:
            return Http404
        if company_phone_number is None or len(company_phone_number) == 0:
            return Http404
        if contact_person_name is None or len(contact_person_name) == 0:
            return Http404
        if contact_person_position is None or len(contact_person_position) == 0:
            return Http404

        try:
            os.makedirs(settings.DOC_GENERATED_DIRS, exist_ok=True)

            if Position.objects.filter(id=contact_person_position).exists():
                position = Position.objects.filter(id=contact_person_position).first()
                if position.translations.filter(language_code=lang).exists():
                    contact_person_position = position.translations.filter(language_code=lang).first().name
                else:
                    contact_person_position = position.translated.name

            if lang == "en":
                file_template = os.path.join(settings.DOC_TEMPLATES_DIRS, "cooperation_agreement_template_en.docx")
                download_file = "Cooperation agreement to provide services on ecommerce platform NextPro.io.pdf"
            else:
                file_template = os.path.join(settings.DOC_TEMPLATES_DIRS, "cooperation_agreement_template_vi.docx")
                download_file = "Thỏa thuận hợp tác cung cấp dịch vụ giao dịch trên sàn giao dịch TMĐT NextPro.io.pdf"
            doc = DocxTemplate(file_template)
            current_date = date.today()
            context = {
                'day': current_date.day,
                'month': current_date.month,
                'year': current_date.year,
                'company_name': company_full_name,
                'company_address': company_address,
                'company_tax_number': company_tax_code,
                'company_phone_number': company_phone_number,
                'contact_person_name': contact_person_name,
                'contact_person_position': contact_person_position
            }
            file_name = str(uuid.uuid4())
            doc_file = os.path.join(settings.DOC_GENERATED_DIRS, file_name + ".docx")
            pdf_file = os.path.join(settings.DOC_GENERATED_DIRS, file_name + ".pdf")

            doc.render(context)
            doc.save(doc_file)
            os.system("lowriter --convert-to pdf " + doc_file + " --outdir " + settings.DOC_GENERATED_DIRS)
            os.remove(doc_file)

            pdf = open(pdf_file, 'rb')
            response = HttpResponse(pdf.read(), content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename=' + download_file
            pdf.close()
            os.remove(pdf_file)

            return response
        except:
            return Http404

def export_students(request):
    # Lọc các user có user_type = 2
    users = User.objects.filter(user_type=2)

    # Tạo workbook và sheet
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Danh sách Người Mua'

    sheet['A1'] = 'Email'
    sheet['B1'] = 'Username'
    sheet['C1'] = 'MSSV'
    sheet['D1'] = 'Tên ngắn'
    sheet['E1'] = 'Ngày sinh'
    sheet['F1'] = 'Nơi sinh'
    sheet['G1'] = 'Lớp'
    sheet['H1'] = 'Khóa học'
    sheet['I1'] = 'Bậc đào tạo'
    sheet['J1'] = 'Loại hình đào tạo'
    sheet['K1'] = 'Ngành'
    sheet['L1'] = 'Giới tính'

    # Định nghĩa kiểu ngày tháng
    date_style = NamedStyle(name="date_style", number_format="DD/MM/YYYY")
    
    # Điền dữ liệu vào các hàng
    for index, user in enumerate(users, start=2):
        sheet[f'A{index}'] = user.email
        sheet[f'B{index}'] = user.username
        sheet[f'C{index}'] = user.mssv
        sheet[f'D{index}'] = user.short_name
        sheet[f'E{index}'] = user.ngay_sinh
        sheet[f'F{index}'] = user.noi_sinh
        sheet[f'G{index}'] = user.lop
        sheet[f'H{index}'] = user.khoa_hoc
        sheet[f'I{index}'] = user.bac_dao_tao
        sheet[f'J{index}'] = user.loai_hinh_dao_tao
        sheet[f'K{index}'] = user.nganh
        sheet[f'L{index}'] = user.gender

        # Áp dụng kiểu định dạng ngày cho cột 'Ngày sinh'
        sheet[f'E{index}'].style = date_style

    # Điều chỉnh chiều rộng các cột (Tùy chọn)
    for col in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in sheet.iter_rows(min_col=col, max_col=col):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Tạo HTTP response để trả về tệp Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=students.xlsx'

    # Lưu workbook vào response
    wb.save(response)

    return response